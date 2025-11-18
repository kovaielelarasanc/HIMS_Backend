from __future__ import annotations
from fastapi import Request, APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import or_
from sqlalchemy.orm import Session
from decimal import Decimal
from typing import List, Optional, Literal
import csv
import io

from app.api.deps import get_db, current_user as auth_current_user
from app.models.user import User
from app.models.pharmacy import (PharmacyMedicine, PharmacySupplier,
                                 PharmacyLocation)
from app.schemas.pharmacy import (MedicineIn, MedicineOut, SupplierIn,
                                  SupplierOut, LocationIn, LocationOut)

router = APIRouter()


# ---------------- Auth helper ----------------
def has_perm(user: User, code: str) -> bool:
    if getattr(user, "is_admin", False):
        return True
    for r in (getattr(user, "roles", None) or []):
        for p in (getattr(r, "permissions", None) or []):
            if getattr(p, "code", None) == code:
                return True
    return False


# =======================================================================
#                               MEDICINES
# =======================================================================

MED_COLUMNS = [
    "code", "name", "generic_name", "form", "strength", "unit", "pack_size",
    "manufacturer", "class_name", "atc_code", "lasa_flag",
    "default_tax_percent", "default_price", "default_mrp", "reorder_level",
    "is_active"
]


def _loose_bool(v, default=True) -> bool:
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ("1", "true", "t", "yes", "y")


def _sample_response(kind: Optional[str]) -> StreamingResponse:
    """
    Build the medicine sample file (CSV or XLSX). Defaults to XLSX when kind is None/invalid.
    Accepts: 'csv' | 'xlsx'
    """
    kind = (kind or "xlsx").lower()

    if kind == "csv":
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=MED_COLUMNS)
        w.writeheader()
        w.writerow({
            "code": "AMOX500",
            "name": "Amoxicillin 500",
            "generic_name": "Amoxicillin",
            "form": "tablet",
            "strength": "500 mg",
            "unit": "tablet",
            "pack_size": 10,
            "manufacturer": "ACME Pharma",
            "class_name": "Antibiotic",
            "atc_code": "J01CA04",
            "lasa_flag": False,
            "default_tax_percent": 12,
            "default_price": 3.50,
            "default_mrp": 5.00,
            "reorder_level": 50,
            "is_active": True,
        })
        data = buf.getvalue().encode("utf-8")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv",
            headers={
                "Content-Disposition":
                "attachment; filename=medicine_import_sample.csv"
            },
        )

    # XLSX default
    try:
        from openpyxl import Workbook
    except Exception:
        return _sample_response("csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "medicines"
    ws.append(MED_COLUMNS)
    ws.append([
        "AMOX500", "Amoxicillin 500", "Amoxicillin", "tablet", "500 mg",
        "tablet", 10, "ACME Pharma", "Antibiotic", "J01CA04", False, 12, 3.50,
        5.00, 50, True
    ])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            "attachment; filename=medicine_import_sample.xlsx"
        },
    )


# ---------- List / Get / CRUD ----------
@router.get("/medicines", response_model=List[MedicineOut])
def list_medicines(
        q: Optional[str] = Query(
            None, description="Search code/name/generic/manufacturer"),
        form: Optional[str] = None,
        class_name: Optional[str] = None,
        is_active: Optional[bool] = Query(None,
                                          description="Filter by active flag"),
        limit: int = Query(200, ge=1, le=1000),
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "pharmacy.view"):
        raise HTTPException(403, "Not permitted")

    qry = db.query(PharmacyMedicine)
    if q:
        like = f"%{q.strip()}%"
        qry = qry.filter(
            or_(
                PharmacyMedicine.code.ilike(like),
                PharmacyMedicine.name.ilike(like),
                PharmacyMedicine.generic_name.ilike(like),
                PharmacyMedicine.manufacturer.ilike(like),
            ))
    if form:
        qry = qry.filter(PharmacyMedicine.form == form)
    if class_name:
        qry = qry.filter(PharmacyMedicine.class_name.ilike(f"%{class_name}%"))
    if is_active is not None:
        qry = qry.filter(PharmacyMedicine.is_active.is_(bool(is_active)))

    return qry.order_by(PharmacyMedicine.name.asc()).limit(limit).all()


@router.get("/medicines/{mid}", response_model=MedicineOut)
def get_medicine(
        mid: int,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "pharmacy.view"):
        raise HTTPException(403, "Not permitted")
    m = db.query(PharmacyMedicine).get(mid)
    if not m:
        raise HTTPException(404, "Not found")
    return m


@router.post("/medicines", response_model=MedicineOut)
def create_medicine(
        payload: MedicineIn,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "pharmacy.masters.manage"):
        raise HTTPException(403, "Not permitted")
    if db.query(PharmacyMedicine).filter(
            PharmacyMedicine.code == payload.code).first():
        raise HTTPException(400, "Code already exists")

    m = PharmacyMedicine(**payload.dict())
    db.add(m)
    db.commit()
    db.refresh(m)
    return m


@router.patch("/medicines/{mid}", response_model=MedicineOut)
@router.put("/medicines/{mid}", response_model=MedicineOut)
def update_medicine(
        mid: int,
        payload: MedicineIn,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "pharmacy.masters.manage"):
        raise HTTPException(403, "Not permitted")

    m = db.query(PharmacyMedicine).get(mid)
    if not m:
        raise HTTPException(404, "Not found")

    if payload.code != m.code and db.query(PharmacyMedicine).filter(
            PharmacyMedicine.code == payload.code).first():
        raise HTTPException(400, "Code already exists")

    for k, v in payload.dict().items():
        setattr(m, k, v)

    db.commit()
    db.refresh(m)
    return m


@router.delete("/medicines/{mid}")
def delete_medicine(
        mid: int,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "pharmacy.masters.manage"):
        raise HTTPException(403, "Not permitted")

    m = db.query(PharmacyMedicine).get(mid)
    if not m:
        raise HTTPException(404, "Not found")

    m.is_active = False  # soft delete
    db.commit()
    return {"ok": True, "message": "Deactivated"}


# ---------- Sample download (supports query & extension styles) ----------
# ---------- Sample download (supports query & extension styles) ----------
# 100% collision-proof sample download endpoints
@router.get("/medicines/samples/template")  # /api/pharmacy/medicines/samples/template?fmt=csv|xlsx
def ph_meds_sample_template(request: Request):
    fmt = (request.query_params.get("fmt")
           or request.query_params.get("format") or "xlsx").lower()
    return _sample_response(fmt)

@router.get("/medicines/samples/template.{ext}")  # /api/pharmacy/medicines/samples/template.csv|xlsx
def ph_meds_sample_template_ext(ext: Literal["csv", "xlsx"]):
    return _sample_response(ext)

@router.get("/medicines/samples/example")  # alias
def ph_meds_sample_example(request: Request):
    fmt = (request.query_params.get("fmt")
           or request.query_params.get("format") or "xlsx").lower()
    return _sample_response(fmt)

@router.get("/medicines/samples/example.{ext}")
def ph_meds_sample_example_ext(ext: Literal["csv", "xlsx"]):
    return _sample_response(ext)



# ---------- Import ----------
def _to_bool(v):
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ("1", "true", "yes", "y", "t")


def _to_dec(v):
    if v is None or str(v).strip() == "":
        return None
    return Decimal(str(v))


def _to_int(v, default=0):
    if v is None or str(v).strip() == "":
        return default
    return int(float(v))


@router.post("/medicines/import")
async def import_medicines(
        request: Request,
        file: UploadFile = File(...),
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "pharmacy.masters.manage"):
        raise HTTPException(403, "Not permitted")

    fmt = (request.query_params.get("fmt")
           or request.query_params.get("format") or "").lower()

    def _str_to_bool(v: Optional[str], default=True) -> bool:
        if v is None:
            return default
        s = str(v).strip().lower()
        return s not in ("0", "false", "no", "off")

    upsert = _str_to_bool(request.query_params.get("upsert"), default=True)

    filename = (file.filename or "").lower()
    content = await file.read()
    rows: List[dict] = []

    use_csv = (fmt == "csv") or filename.endswith(".csv")
    use_xlsx = (fmt
                == "xlsx") or filename.endswith(".xlsx") or (fmt == ""
                                                             and not use_csv)

    if use_csv:
        buf = io.StringIO(content.decode("utf-8-sig"))
        rdr = csv.DictReader(buf)
        for r in rdr:
            rows.append(r)
    elif use_xlsx:
        try:
            from openpyxl import load_workbook
        except Exception:
            raise HTTPException(
                400,
                "Install 'openpyxl' to import .xlsx files or upload a .csv")
        bio = io.BytesIO(content)
        wb = load_workbook(bio, read_only=True)
        ws = wb.active
        headers = [(str(c.value).strip() if c.value is not None else "")
                   for c in next(ws.rows)]
        missing = [h for h in MED_COLUMNS if h not in headers]
        if missing:
            raise HTTPException(400, f"Missing columns: {missing}")
        idx = {h: headers.index(h) for h in MED_COLUMNS}
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({
                h: row[idx[h]] if idx[h] < len(row) else None
                for h in MED_COLUMNS
            })
    else:
        raise HTTPException(400, "Unsupported file (use .xlsx or .csv)")

    inserted = updated = skipped = 0
    errors: List[dict] = []

    for i, r in enumerate(rows, start=2):
        try:
            code = (r.get("code") or "").strip()
            name = (r.get("name") or "").strip()
            form = (r.get("form") or "").strip()
            if not code or not name or not form:
                skipped += 1
                continue

            payload = {
                "code":
                code,
                "name":
                name,
                "generic_name": (r.get("generic_name") or "").strip(),
                "form":
                form,
                "strength": (r.get("strength") or "").strip(),
                "unit": (r.get("unit") or "unit").strip(),
                "pack_size":
                _to_int(r.get("pack_size"), 1),
                "manufacturer": (r.get("manufacturer") or "").strip(),
                "class_name": (r.get("class_name") or "").strip(),
                "atc_code": (r.get("atc_code") or "").strip(),
                "lasa_flag":
                _to_bool(r.get("lasa_flag")),
                "default_tax_percent":
                _to_dec(r.get("default_tax_percent")),
                "default_price":
                _to_dec(r.get("default_price")),
                "default_mrp":
                _to_dec(r.get("default_mrp")),
                "reorder_level":
                _to_int(r.get("reorder_level"), 0),
                "is_active":
                _to_bool(r.get("is_active"))
                if r.get("is_active") is not None else True,
            }

            existing = db.query(PharmacyMedicine).filter(
                PharmacyMedicine.code == code).first()
            if existing:
                if not upsert:
                    skipped += 1
                    continue
                for k, v in payload.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                db.add(PharmacyMedicine(**payload))
                inserted += 1

        except Exception as ex:
            errors.append({"row": i, "error": str(ex)})

    db.commit()
    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "total_rows": len(rows),
    }


# =======================================================================
#                               SUPPLIERS
# =======================================================================


@router.get("/suppliers", response_model=List[SupplierOut])
def list_suppliers(
        q: Optional[str] = None,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "pharmacy.view"):
        raise HTTPException(403, "Not permitted")
    qry = db.query(PharmacySupplier)
    if q:
        like = f"%{q.strip()}%"
        qry = qry.filter(
            or_(
                PharmacySupplier.name.ilike(like),
                PharmacySupplier.phone.ilike(like),
                PharmacySupplier.email.ilike(like),
                PharmacySupplier.gstin.ilike(like),
            ))
    return qry.order_by(PharmacySupplier.name.asc()).all()


@router.post("/suppliers", response_model=SupplierOut)
def create_supplier(
        payload: SupplierIn,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "pharmacy.masters.manage"):
        raise HTTPException(403, "Not permitted")
    s = PharmacySupplier(**payload.dict())
    db.add(s)
    db.commit()
    db.refresh(s)
    return s


@router.patch("/suppliers/{sid}", response_model=SupplierOut)
@router.put("/suppliers/{sid}", response_model=SupplierOut)
def update_supplier(
        sid: int,
        payload: SupplierIn,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "pharmacy.masters.manage"):
        raise HTTPException(403, "Not permitted")
    s = db.query(PharmacySupplier).get(sid)
    if not s:
        raise HTTPException(404, "Not found")
    for k, v in payload.dict().items():
        setattr(s, k, v)
    db.commit()
    db.refresh(s)
    return s


@router.delete("/suppliers/{sid}")
def delete_supplier(
        sid: int,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "pharmacy.masters.manage"):
        raise HTTPException(403, "Not permitted")
    s = db.query(PharmacySupplier).get(sid)
    if not s:
        raise HTTPException(404, "Not found")
    db.delete(s)
    db.commit()
    return {"ok": True}


# =======================================================================
#                               LOCATIONS
# =======================================================================


@router.get("/locations", response_model=List[LocationOut])
def list_locations(
        q: Optional[str] = None,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "pharmacy.view"):
        raise HTTPException(403, "Not permitted")
    qry = db.query(PharmacyLocation)
    if q:
        like = f"%{q.strip()}%"
        qry = qry.filter(
            or_(
                PharmacyLocation.code.ilike(like),
                PharmacyLocation.name.ilike(like),
            ))
    return qry.order_by(PharmacyLocation.code.asc()).all()


@router.post("/locations", response_model=LocationOut)
def create_location(
        payload: LocationIn,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "pharmacy.masters.manage"):
        raise HTTPException(403, "Not permitted")
    if db.query(PharmacyLocation).filter(
            PharmacyLocation.code == payload.code).first():
        raise HTTPException(400, "Code already exists")
    loc = PharmacyLocation(**payload.dict())
    db.add(loc)
    db.commit()
    db.refresh(loc)
    return loc


@router.patch("/locations/{lid}", response_model=LocationOut)
@router.put("/locations/{lid}", response_model=LocationOut)
def update_location(
        lid: int,
        payload: LocationIn,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "pharmacy.masters.manage"):
        raise HTTPException(403, "Not permitted")
    loc = db.query(PharmacyLocation).get(lid)
    if not loc:
        raise HTTPException(404, "Not found")
    if payload.code != loc.code and db.query(PharmacyLocation).filter(
            PharmacyLocation.code == payload.code).first():
        raise HTTPException(400, "Code already exists")
    for k, v in payload.dict().items():
        setattr(loc, k, v)
    db.commit()
    db.refresh(loc)
    return loc


@router.delete("/locations/{lid}")
def delete_location(
        lid: int,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "pharmacy.masters.manage"):
        raise HTTPException(403, "Not permitted")
    loc = db.query(PharmacyLocation).get(lid)
    if not loc:
        raise HTTPException(404, "Not found")
    db.delete(loc)
    db.commit()
    return {"ok": True}
