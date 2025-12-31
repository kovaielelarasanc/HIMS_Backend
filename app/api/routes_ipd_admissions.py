# FILE: app/api/routes_ipd_admissions.py
from __future__ import annotations

import io
from datetime import datetime, date, time
from typing import Optional

from fastapi import APIRouter, Depends, Query, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import func, or_

from starlette.responses import StreamingResponse

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.api.deps import get_db, current_user
from app.models.ipd import IpdAdmission, IpdBed, IpdRoom, IpdWard
from app.models.patient import Patient
from app.models.user import User  # must exist in your project
from app.schemas.ipd_admissions import IpdAdmissionListOut, IpdAdmissionListItem

router = APIRouter(prefix="/ipds", tags=["IPD Admissions"])


# --------------------------
# Permissions
# --------------------------
def _need_any(user: User, codes: list[str]) -> None:
    if getattr(user, "is_admin", False):
        return
    have = {p.code for r in (user.roles or []) for p in (getattr(r, "permissions", []) or [])}
    if have.intersection(set(codes)):
        return
    raise HTTPException(status_code=403, detail="Not permitted")


def _safe_trim(s: str) -> str:
    return (s or "").strip()


def _patient_name_expr():
    # MySQL-safe "First Last" with NULL handling
    return func.trim(
        func.concat(
            func.coalesce(Patient.first_name, ""),
            " ",
            func.coalesce(Patient.last_name, ""),
        )
    )


def _doctor_name_expr():
    return func.coalesce(User.name, "")


def _parse_date_or_datetime(v: Optional[str], *, end_of_day: bool = False) -> Optional[datetime]:
    """
    Accepts:
      - "YYYY-MM-DD"
      - "YYYY-MM-DDTHH:MM"
      - "YYYY-MM-DD HH:MM:SS"
      - ISO datetime strings
    Returns datetime (IST conversion NOT done here; we filter DB stored datetime).
    """
    if not v:
        return None
    s = v.strip()
    if not s:
        return None

    # date-only
    try:
        if len(s) == 10 and s[4] == "-" and s[7] == "-":
            d = date.fromisoformat(s)
            if end_of_day:
                return datetime.combine(d, time.max)
            return datetime.combine(d, time.min)
    except Exception:
        pass

    # datetime string
    try:
        # allow " " instead of "T"
        s2 = s.replace(" ", "T")
        return datetime.fromisoformat(s2)
    except Exception:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid date/datetime format: {v}. Use YYYY-MM-DD or ISO datetime.",
        )


def _build_query(
    db: Session,
    *,
    q: str = "",
    status: str = "",
    from_admit: Optional[datetime] = None,
    to_admit: Optional[datetime] = None,
    doctor_user_id: Optional[int] = None,
    department_id: Optional[int] = None,
):
    PNAME = _patient_name_expr()
    DNAME = _doctor_name_expr()

    qry = (
        db.query(
            IpdAdmission.id.label("id"),
            IpdAdmission.admission_code.label("admission_code"),
            IpdAdmission.status.label("status"),
            IpdAdmission.admitted_at.label("admitted_at"),
            IpdAdmission.discharge_at.label("discharge_at"),
            Patient.id.label("patient_id"),
            func.coalesce(Patient.uhid, "—").label("uhid"),
            PNAME.label("patient_name"),
            User.id.label("doctor_user_id"),
            DNAME.label("doctor_name"),
            IpdWard.name.label("ward_name"),
            IpdRoom.number.label("room_number"),
            IpdBed.code.label("bed_code"),
        )
        .outerjoin(Patient, Patient.id == IpdAdmission.patient_id)
        .outerjoin(User, User.id == IpdAdmission.practitioner_user_id)
        .outerjoin(IpdBed, IpdBed.id == IpdAdmission.current_bed_id)
        .outerjoin(IpdRoom, IpdRoom.id == IpdBed.room_id)
        .outerjoin(IpdWard, IpdWard.id == IpdRoom.ward_id)
    )

    # --------------------------
    # Filters
    # --------------------------
    if status:
        qry = qry.filter(IpdAdmission.status == status)

    if from_admit:
        qry = qry.filter(IpdAdmission.admitted_at >= from_admit)

    if to_admit:
        qry = qry.filter(IpdAdmission.admitted_at <= to_admit)

    # doctor filter (by id, not name)
    if doctor_user_id:
        qry = qry.filter(IpdAdmission.practitioner_user_id == doctor_user_id)

    # department filter (safe if your model has department_id)
    dep_col = getattr(IpdAdmission, "department_id", None)
    if department_id and dep_col is not None:
        qry = qry.filter(dep_col == department_id)

    # unified search only
    if q:
        s = q.strip()
        qry = qry.filter(
            or_(
                Patient.uhid.ilike(f"%{s}%"),
                Patient.first_name.ilike(f"%{s}%"),
                Patient.last_name.ilike(f"%{s}%"),
                PNAME.ilike(f"%{s}%"),
                DNAME.ilike(f"%{s}%"),
                IpdAdmission.admission_code.ilike(f"%{s}%"),
            )
        )

    qry = qry.order_by(IpdAdmission.admitted_at.desc(), IpdAdmission.id.desc())
    return qry


def _to_item(row) -> IpdAdmissionListItem:
    admission_code = row.admission_code or f"IP-{int(row.id):06d}"
    doctor_name = _safe_trim(row.doctor_name) or "—"

    return IpdAdmissionListItem(
        id=int(row.id),
        admission_code=admission_code,
        patient_id=int(row.patient_id) if row.patient_id else 0,
        patient_name=_safe_trim(row.patient_name) or "—",
        uhid=_safe_trim(row.uhid) or "—",
        doctor_user_id=int(row.doctor_user_id) if row.doctor_user_id else None,
        doctor_name=doctor_name,
        ward_name=_safe_trim(row.ward_name) or None,
        room_number=_safe_trim(row.room_number) or None,
        bed_code=_safe_trim(row.bed_code) or None,
        status=_safe_trim(row.status) or "admitted",
        admitted_at=row.admitted_at,
        discharge_at=row.discharge_at,
    )


@router.get("/admissions", response_model=IpdAdmissionListOut)
def list_ipd_admissions(
    q: str = Query("", description="Search across UHID / Patient / Doctor / Admission Code"),
    status: str = Query("", description="Filter by status"),
    from_admit: Optional[str] = Query(None, description="From admit date (YYYY-MM-DD)"),
    to_admit: Optional[str] = Query(None, description="To admit date (YYYY-MM-DD)"),
    doctor_user_id: Optional[int] = Query(None, description="Doctor user id"),
    department_id: Optional[int] = Query(None, description="Department id (if admission has department_id)"),
    limit: int = Query(30, ge=1, le=200),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    user=Depends(current_user),
):
    _need_any(user, ["ipd.admissions.view", "ipd.view", "ipd.manage"])

    fdt = _parse_date_or_datetime(from_admit, end_of_day=False)
    tdt = _parse_date_or_datetime(to_admit, end_of_day=True)

    base = _build_query(
        db,
        q=q,
        status=status,
        from_admit=fdt,
        to_admit=tdt,
        doctor_user_id=doctor_user_id,
        department_id=department_id,
    )

    total = (
        base.order_by(None)
        .with_entities(func.count(func.distinct(IpdAdmission.id)))
        .scalar()
        or 0
    )

    rows = base.limit(limit).offset(offset).all()
    items = [_to_item(r) for r in rows]
    return IpdAdmissionListOut(items=items, total=int(total), limit=limit, offset=offset)


@router.get("/admissions/export")
def export_ipd_admissions_excel(
    q: str = Query(""),
    status: str = Query(""),
    from_admit: Optional[str] = Query(None),
    to_admit: Optional[str] = Query(None),
    doctor_user_id: Optional[int] = Query(None),
    department_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    user=Depends(current_user),
):
    _need_any(user, ["ipd.admissions.view", "ipd.view", "ipd.manage"])

    fdt = _parse_date_or_datetime(from_admit, end_of_day=False)
    tdt = _parse_date_or_datetime(to_admit, end_of_day=True)

    qry = _build_query(
        db,
        q=q,
        status=status,
        from_admit=fdt,
        to_admit=tdt,
        doctor_user_id=doctor_user_id,
        department_id=department_id,
    )

    rows = qry.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "IPD Admissions"

    headers = [
        "Admission Code",
        "Patient Name",
        "UHID",
        "Doctor Name",
        "Ward",
        "Room",
        "Bed",
        "Status",
        "Admitted At",
        "Discharge At",
    ]
    ws.append(headers)

    for r in rows:
        item = _to_item(r)
        ws.append(
            [
                item.admission_code,
                item.patient_name,
                item.uhid,
                item.doctor_name,
                item.ward_name or "",
                item.room_number or "",
                item.bed_code or "",
                item.status,
                item.admitted_at.isoformat(sep=" ") if item.admitted_at else "",
                item.discharge_at.isoformat(sep=" ") if item.discharge_at else "",
            ]
        )

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"ipd_admissions_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
