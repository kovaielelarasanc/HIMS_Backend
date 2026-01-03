# FILE: app/api/routes_pharmacy_stock_alerts.py
from __future__ import annotations

from io import BytesIO
from datetime import date as dt_date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, func

from app.utils.resp import ok, err

# ✅ Adjust these imports to your project structure if needed
from app.api.deps import get_db, current_user
from app.models.user import User

from app.models.pharmacy_inventory import (
    InventoryLocation,
    InventoryItem,
    ItemBatch,
)

from app.schemas.pharmacy_stock_alerts import (
    AlertType,
    ReportType,
)

from app.services.pharmacy_stock_alerts import (
    get_dashboard,
    list_alerts,
    list_item_batches,
    build_report_rows,
)

# PDF + Excel helpers
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


router = APIRouter(prefix="/pharmacy/stock", tags=["Pharmacy Stock & Alerts"])


# -------------------------
# Permission helper (robust)
# -------------------------
def _need_any(user: User, perms: List[str]) -> None:
    """
    Robust permission check:
    - Passes if superuser/admin flags exist
    - Else checks common permission storage patterns:
        user.has_permission("x")
        "x" in user.permissions
    If your project already has a standard _need_any, you can remove this and import yours.
    """
    try:
        if getattr(user, "is_superuser", False) or getattr(user, "is_admin", False):
            return
        has_fn = getattr(user, "has_permission", None)
        if callable(has_fn):
            for p in perms:
                if has_fn(p):
                    return
        perms_list = getattr(user, "permissions", None) or []
        for p in perms:
            if p in perms_list:
                return
    except Exception:
        # If user object doesn't support perms, do not hard-fail here.
        return

    # If we can check perms and none match, block
    # (If you prefer open access, comment this out.)
    from fastapi import HTTPException
    raise HTTPException(status_code=403, detail="Not permitted")


# -------------------------
# Small helpers
# -------------------------
def _resolve_pharmacy_locations(db: Session, location_id: Optional[int]) -> List[InventoryLocation]:
    q = db.query(InventoryLocation).filter(
        InventoryLocation.is_active.is_(True),
        InventoryLocation.is_pharmacy.is_(True),
    )
    if location_id:
        q = q.filter(InventoryLocation.id == location_id)
    return q.order_by(InventoryLocation.name.asc()).all()


def _order_nulls_last(col):
    # MySQL safe ordering (no nullslast())
    return (col.is_(None).asc(), col.asc())


def _apply_item_filters(
    q,
    item_type: Optional[str],
    schedule_code: Optional[str],
    supplier_id: Optional[int],
    category_id: Optional[int],
):
    if item_type:
        q = q.filter(InventoryItem.item_type == item_type)
    if schedule_code:
        q = q.filter(InventoryItem.schedule_code == schedule_code)
    if supplier_id:
        q = q.filter(InventoryItem.default_supplier_id == supplier_id)
    # optional if you have category_id
    if category_id and hasattr(InventoryItem, "category_id"):
        q = q.filter(getattr(InventoryItem, "category_id") == category_id)
    return q


def _safe_int(v: Any, default: int = 0) -> int:
    try:
        return int(v)
    except Exception:
        return default


# ============================================================
# 1) DASHBOARD SUMMARY
# ============================================================
@router.get("/alerts/summary")
def stock_alerts_summary(
    db: Session = Depends(get_db),
    user: User = Depends(current_user),

    # filters
    location_id: Optional[int] = Query(None),
    item_type: Optional[str] = Query(None),
    schedule_code: Optional[str] = Query(None),
    supplier_id: Optional[int] = Query(None),

    days_near_expiry: Optional[int] = Query(None),
    non_moving_days: int = Query(60),
    fast_moving_days: int = Query(30),
    consumption_days: int = Query(30),
    lead_time_days: int = Query(7),

    high_value_expiry_threshold: Decimal = Query(Decimal("0")),
    preview_limit: int = Query(25, ge=5, le=200),
):
    try:
        _need_any(user, ["pharmacy.stock.view", "pharmacy.inventory.view", "pharmacy.view"])

        data = get_dashboard(
            db=db,
            location_id=location_id,
            item_type=item_type,
            schedule_code=schedule_code,
            supplier_id=supplier_id,
            days_near_expiry=days_near_expiry,
            non_moving_days=non_moving_days,
            fast_moving_days=fast_moving_days,
            consumption_days=consumption_days,
            lead_time_days=lead_time_days,
            high_value_expiry_threshold=high_value_expiry_threshold,
            preview_limit=preview_limit,
        )
        return ok(data)

    except Exception as e:
        return err(f"Failed to load stock alerts summary: {str(e)}", status_code=500)


# ============================================================
# 2) ALERTS LIST (Action table backend)
# ============================================================
@router.get("/alerts/list")
def stock_alerts_list(
    db: Session = Depends(get_db),
    user: User = Depends(current_user),

    # accept both alert_type and type (frontend friendliness)
    alert_type: Optional[AlertType] = Query(None),
    type_: Optional[AlertType] = Query(None, alias="type"),

    location_id: Optional[int] = Query(None),
    item_type: Optional[str] = Query(None),
    schedule_code: Optional[str] = Query(None),
    supplier_id: Optional[int] = Query(None),

    days_near_expiry: int = Query(90),
    non_moving_days: int = Query(60),
    consumption_days: int = Query(30),
    lead_time_days: int = Query(7),
    high_value_expiry_threshold: Decimal = Query(Decimal("0")),

    include_batches: bool = Query(True),
    limit: int = Query(200, ge=1, le=2000),
    offset: int = Query(0, ge=0),
):
    try:
        _need_any(user, ["pharmacy.stock.view", "pharmacy.inventory.view", "pharmacy.view"])

        a_type = alert_type or type_
        if not a_type:
            return err("alert_type is required", status_code=400)

        rows = list_alerts(
            db=db,
            alert_type=a_type,
            location_id=location_id,
            item_type=item_type,
            schedule_code=schedule_code,
            supplier_id=supplier_id,
            days_near_expiry=days_near_expiry,
            non_moving_days=non_moving_days,
            consumption_days=consumption_days,
            lead_time_days=lead_time_days,
            high_value_expiry_threshold=high_value_expiry_threshold,
            include_batches=include_batches,
            limit=limit,
            offset=offset,
        )
        return ok(rows)

    except Exception as e:
        return err(f"Failed to load alerts list: {str(e)}", status_code=500)


# ============================================================
# 3) ITEM -> ALL BATCHES (mandatory batch-wise)
# ============================================================
@router.get("/items/{item_id}/batches")
def item_batches(
    item_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(current_user),
    location_id: Optional[int] = Query(None),
):
    try:
        _need_any(user, ["pharmacy.stock.view", "pharmacy.inventory.view", "pharmacy.view"])

        rows = list_item_batches(db=db, item_id=item_id, location_id=location_id)
        return ok(rows)

    except Exception as e:
        return err(f"Failed to load item batches: {str(e)}", status_code=500)


# ============================================================
# 4) EXPORT REPORTS (Excel + PDF)
# ============================================================
def _export_xlsx(rows: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    headers: List[str] = []
    if rows:
        headers = list(rows[0].keys())
    else:
        headers = ["Report", "Message"]

    ws.append(headers)

    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    # autosize columns
    for i, h in enumerate(headers, start=1):
        col = get_column_letter(i)
        max_len = max(len(str(h)), 10)
        for cell in ws[col]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col].width = min(max_len + 2, 55)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _export_pdf(rows: List[Dict[str, Any]], title: str = "Stock Report") -> bytes:
    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    w, h = A4

    x = 12 * mm
    y = h - 15 * mm

    c.setFont("Helvetica-Bold", 11)
    c.drawString(x, y, title)
    y -= 8 * mm

    c.setFont("Helvetica", 8)

    # columns: pick a stable subset for readability
    preferred = [
        "Location", "Item Name", "Batch No", "Expiry Date", "Qty",
        "Unit Cost", "MRP", "Value (Purchase)", "Value (MRP)"
    ]
    if not rows:
        c.drawString(x, y, "No data")
        c.showPage()
        c.save()
        return bio.getvalue()

    cols = [k for k in preferred if k in rows[0].keys()]
    if not cols:
        cols = list(rows[0].keys())[:8]

    # simple fixed widths
    widths = [30*mm, 45*mm, 25*mm, 20*mm, 15*mm, 18*mm, 18*mm, 22*mm, 22*mm][:len(cols)]

    def draw_row(vals: List[str], yy: float, bold: bool = False):
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 8)
        xx = x
        for i, v in enumerate(vals):
            c.drawString(xx, yy, (v or "")[:35])
            xx += widths[i] if i < len(widths) else 25*mm

    # header
    draw_row(cols, y, bold=True)
    y -= 5 * mm

    for r in rows[:5000]:  # hard safety limit
        if y < 12 * mm:
            c.showPage()
            y = h - 15 * mm
            c.setFont("Helvetica", 8)
            draw_row(cols, y, bold=True)
            y -= 5 * mm
        draw_row([str(r.get(k, "")) for k in cols], y, bold=False)
        y -= 4.5 * mm

    c.showPage()
    c.save()
    return bio.getvalue()


@router.get("/alerts/export")
def export_stock_reports(
    db: Session = Depends(get_db),
    user: User = Depends(current_user),

    report_type: Optional[ReportType] = Query(None),
    type_: Optional[ReportType] = Query(None, alias="type"),

    format: str = Query("xlsx", regex="^(xlsx|pdf)$"),

    location_id: Optional[int] = Query(None),
    item_type: Optional[str] = Query(None),
    schedule_code: Optional[str] = Query(None),
    supplier_id: Optional[int] = Query(None),

    days_near_expiry: int = Query(90),
    non_moving_days: int = Query(60),
    consumption_days: int = Query(30),
    lead_time_days: int = Query(7),

    limit: int = Query(20000, ge=1, le=50000),
    offset: int = Query(0, ge=0),
):
    try:
        _need_any(user, ["pharmacy.stock.view", "pharmacy.inventory.view", "pharmacy.view"])

        rtype = report_type or type_
        if not rtype:
            return err("report_type is required", status_code=400)

        rows = build_report_rows(
            db=db,
            report_type=rtype,
            location_id=location_id,
            item_type=item_type,
            schedule_code=schedule_code,
            supplier_id=supplier_id,
            days_near_expiry=days_near_expiry,
            non_moving_days=non_moving_days,
            consumption_days=consumption_days,
            lead_time_days=lead_time_days,
            limit=limit,
            offset=offset,
        )

        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename_base = f"{rtype.value.lower()}_{ts}"

        if format == "pdf":
            pdf_bytes = _export_pdf(rows, title=f"Pharmacy Stock Report - {rtype.value}")
            return StreamingResponse(
                BytesIO(pdf_bytes),
                media_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="{filename_base}.pdf"'},
            )

        xlsx_bytes = _export_xlsx(rows)
        return StreamingResponse(
            BytesIO(xlsx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
        )

    except Exception as e:
        return err(f"Failed to export report: {str(e)}", status_code=500)


# ============================================================
# 5) STOCK LIST (ALL locations default, batch-wise price)
#    Used by frontend: getStockSummary()
# ============================================================
@router.get("/summary")
def stock_summary_batchwise(
    db: Session = Depends(get_db),
    user: User = Depends(current_user),

    # IMPORTANT: if not provided -> ALL pharmacy locations
    location_id: Optional[int] = Query(None),

    q: Optional[str] = Query(None, description="Search by item name/code/batch"),
    item_type: Optional[str] = Query(None),
    schedule_code: Optional[str] = Query(None),
    supplier_id: Optional[int] = Query(None),
    category_id: Optional[int] = Query(None),

    include_zero: bool = Query(False),
    only_saleable: bool = Query(False),

    limit: int = Query(200, ge=1, le=2000),
    offset: int = Query(0, ge=0),
):
    try:
        _need_any(user, ["pharmacy.stock.view", "pharmacy.inventory.view", "pharmacy.view"])

        locs = _resolve_pharmacy_locations(db, location_id)
        loc_ids = [l.id for l in locs]
        if not loc_ids:
            return ok({"total": 0, "rows": [], "totals": {"value_purchase": "0", "value_mrp": "0"}})

        base = (
            db.query(
                ItemBatch.id.label("batch_id"),
                ItemBatch.location_id.label("location_id"),
                InventoryLocation.name.label("location_name"),

                ItemBatch.item_id.label("item_id"),
                InventoryItem.code.label("item_code"),
                InventoryItem.name.label("item_name"),
                InventoryItem.schedule_code.label("schedule_code"),
                InventoryItem.item_type.label("item_type"),
                InventoryItem.default_supplier_id.label("supplier_id"),

                ItemBatch.batch_no.label("batch_no"),
                ItemBatch.expiry_date.label("expiry_date"),
                ItemBatch.current_qty.label("qty"),
                ItemBatch.unit_cost.label("unit_cost"),
                ItemBatch.mrp.label("mrp"),
                (ItemBatch.current_qty * ItemBatch.unit_cost).label("value_purchase"),
                (ItemBatch.current_qty * ItemBatch.mrp).label("value_mrp"),
                ItemBatch.is_saleable.label("is_saleable"),
                ItemBatch.status.label("status"),
            )
            .join(InventoryItem, InventoryItem.id == ItemBatch.item_id)
            .join(InventoryLocation, InventoryLocation.id == ItemBatch.location_id)
            .filter(
                ItemBatch.location_id.in_(loc_ids),
                ItemBatch.is_active.is_(True),
                InventoryItem.is_active.is_(True),
            )
        )

        if not include_zero:
            base = base.filter(ItemBatch.current_qty != 0)

        if only_saleable:
            base = base.filter(ItemBatch.is_saleable.is_(True))

        base = _apply_item_filters(base, item_type, schedule_code, supplier_id, category_id)

        if q:
            like = f"%{q.strip()}%"
            base = base.filter(
                or_(
                    InventoryItem.name.ilike(like),
                    InventoryItem.code.ilike(like),
                    ItemBatch.batch_no.ilike(like),
                )
            )

        total = base.with_entities(func.count()).scalar() or 0

        rows = (
            base.order_by(
                InventoryLocation.name.asc(),
                InventoryItem.name.asc(),
                *_order_nulls_last(ItemBatch.expiry_date),
                ItemBatch.batch_no.asc(),
            )
            .offset(offset)
            .limit(limit)
            .all()
        )

        out_rows: List[Dict[str, Any]] = []
        for r in rows:
            out_rows.append(
                {
                    "batch_id": _safe_int(r.batch_id),
                    "location_id": _safe_int(r.location_id),
                    "location_name": str(r.location_name),

                    "item_id": _safe_int(r.item_id),
                    "item_code": str(r.item_code),
                    "item_name": str(r.item_name),
                    "item_type": str(r.item_type or ""),
                    "schedule_code": str(r.schedule_code or ""),
                    "supplier_id": r.supplier_id,

                    "batch_no": str(r.batch_no),
                    "expiry_date": r.expiry_date,
                    "qty": r.qty,
                    "unit_cost": r.unit_cost,
                    "mrp": r.mrp,
                    "value_purchase": r.value_purchase,
                    "value_mrp": r.value_mrp,
                    "is_saleable": bool(r.is_saleable),
                    "status": str(r.status or ""),
                }
            )

        # totals (for current filter)
        totals_row = base.with_entities(
            func.coalesce(func.sum(ItemBatch.current_qty * ItemBatch.unit_cost), 0),
            func.coalesce(func.sum(ItemBatch.current_qty * ItemBatch.mrp), 0),
        ).one()

        totals = {
            "value_purchase": totals_row[0],
            "value_mrp": totals_row[1],
        }

        return ok({"total": int(total), "rows": out_rows, "totals": totals})

    except Exception as e:
        return err(f"Failed to load stock summary: {str(e)}", status_code=500)


# ============================================================
# 6) QUARANTINE LIST (batch-wise)
#    Used by frontend: getQuarantineBatches()
# ============================================================
@router.get("/quarantine")
def quarantine_batches(
    db: Session = Depends(get_db),
    user: User = Depends(current_user),

    # default ALL locations
    location_id: Optional[int] = Query(None),

    q: Optional[str] = Query(None),
    item_type: Optional[str] = Query(None),
    schedule_code: Optional[str] = Query(None),
    supplier_id: Optional[int] = Query(None),
    category_id: Optional[int] = Query(None),

    include_expired: bool = Query(True),

    limit: int = Query(200, ge=1, le=2000),
    offset: int = Query(0, ge=0),
):
    try:
        _need_any(user, ["pharmacy.stock.view", "pharmacy.inventory.view", "pharmacy.view"])

        today = dt_date.today()
        locs = _resolve_pharmacy_locations(db, location_id)
        loc_ids = [l.id for l in locs]
        if not loc_ids:
            return ok({"total": 0, "rows": []})

        quarantine_cond = or_(
            ItemBatch.is_saleable.is_(False),
            ItemBatch.status.in_(["QUARANTINE", "HOLD", "BLOCKED"]),
        )
        if include_expired:
            quarantine_cond = or_(quarantine_cond, and_(ItemBatch.expiry_date.isnot(None), ItemBatch.expiry_date < today))

        base = (
            db.query(
                ItemBatch.id.label("batch_id"),
                ItemBatch.location_id.label("location_id"),
                InventoryLocation.name.label("location_name"),

                ItemBatch.item_id.label("item_id"),
                InventoryItem.code.label("item_code"),
                InventoryItem.name.label("item_name"),
                InventoryItem.schedule_code.label("schedule_code"),
                InventoryItem.item_type.label("item_type"),
                InventoryItem.default_supplier_id.label("supplier_id"),

                ItemBatch.batch_no.label("batch_no"),
                ItemBatch.expiry_date.label("expiry_date"),
                ItemBatch.current_qty.label("qty"),
                ItemBatch.unit_cost.label("unit_cost"),
                ItemBatch.mrp.label("mrp"),
                (ItemBatch.current_qty * ItemBatch.unit_cost).label("value_purchase"),
                (ItemBatch.current_qty * ItemBatch.mrp).label("value_mrp"),
                ItemBatch.is_saleable.label("is_saleable"),
                ItemBatch.status.label("status"),
            )
            .join(InventoryItem, InventoryItem.id == ItemBatch.item_id)
            .join(InventoryLocation, InventoryLocation.id == ItemBatch.location_id)
            .filter(
                ItemBatch.location_id.in_(loc_ids),
                ItemBatch.is_active.is_(True),
                InventoryItem.is_active.is_(True),
                quarantine_cond,
                ItemBatch.current_qty != 0,
            )
        )

        base = _apply_item_filters(base, item_type, schedule_code, supplier_id, category_id)

        if q:
            like = f"%{q.strip()}%"
            base = base.filter(
                or_(
                    InventoryItem.name.ilike(like),
                    InventoryItem.code.ilike(like),
                    ItemBatch.batch_no.ilike(like),
                )
            )

        total = base.with_entities(func.count()).scalar() or 0

        rows = (
            base.order_by(
                InventoryLocation.name.asc(),
                InventoryItem.name.asc(),
                *_order_nulls_last(ItemBatch.expiry_date),
                ItemBatch.batch_no.asc(),
            )
            .offset(offset)
            .limit(limit)
            .all()
        )

        out_rows: List[Dict[str, Any]] = []
        for r in rows:
            out_rows.append(
                {
                    "batch_id": _safe_int(r.batch_id),
                    "location_id": _safe_int(r.location_id),
                    "location_name": str(r.location_name),

                    "item_id": _safe_int(r.item_id),
                    "item_code": str(r.item_code),
                    "item_name": str(r.item_name),
                    "item_type": str(r.item_type or ""),
                    "schedule_code": str(r.schedule_code or ""),
                    "supplier_id": r.supplier_id,

                    "batch_no": str(r.batch_no),
                    "expiry_date": r.expiry_date,
                    "qty": r.qty,
                    "unit_cost": r.unit_cost,
                    "mrp": r.mrp,
                    "value_purchase": r.value_purchase,
                    "value_mrp": r.value_mrp,
                    "is_saleable": bool(r.is_saleable),
                    "status": str(r.status or ""),
                }
            )

        return ok({"total": int(total), "rows": out_rows})

    except Exception as e:
        return err(f"Failed to load quarantine batches: {str(e)}", status_code=500)
