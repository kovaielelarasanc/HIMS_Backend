# FILE: app/api/routes_patients.py
from __future__ import annotations

import os
import io
import base64
import calendar
from pathlib import Path
from typing import List, Optional, Dict, Any
from datetime import date, datetime, timedelta

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    UploadFile,
    File,
    Form,
    Body,
    Query,
    Request,
)
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
# type: ignore
from sqlalchemy.exc import IntegrityError

from openpyxl import Workbook  # ensure openpyxl in requirements.txt
from openpyxl.utils import get_column_letter

from app.api.deps import get_db, current_user as auth_current_user
from app.core.config import settings
from app.models.user import User
from app.models.patient import (
    Patient,
    PatientAddress,
    PatientDocument,
    PatientConsent,
    PatientType,
)
from app.models.payer import Payer, Tpa, CreditPlan
from app.schemas.patient import (
    PatientCreate,
    PatientUpdate,
    PatientOut,
    AddressIn,
    AddressOut,
    DocumentOut,
    ConsentIn,
    ConsentOut,
)
from app.services.audit_logger import log_audit  # adjust path if your pkg is `services`

router = APIRouter()

# --------- utils ----------

UPLOAD_DIR = Path(settings.STORAGE_DIR).joinpath("patient_docs")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def make_uhid(id_num: int) -> str:
    return f"NH-{id_num:06d}"


def has_perm(user: User, code: str) -> bool:
    if user.is_admin:
        return True
    for r in user.roles:
        for p in r.permissions:
            if p.code == code:
                return True
    return False


def get_request_meta(request: Optional[Request]) -> Dict[str, Optional[str]]:
    if request is None:
        return {"ip": None, "ua": None}
    ip = request.client.host if request.client else None
    ua = request.headers.get("user-agent")
    return {"ip": ip, "ua": ua}


def instance_to_audit_dict(obj: Any) -> Dict[str, Any]:
    """
    Convert SQLAlchemy model instance to a JSON-serializable dict
    based on its columns.
    """
    if obj is None:
        return {}
    data: Dict[str, Any] = {}
    for col in obj.__table__.columns:  # type: ignore[attr-defined]
        val = getattr(obj, col.name)
        if isinstance(val, (date, datetime)):
            val = val.isoformat()
        data[col.name] = val
    return data


def calc_age(dob: Optional[date]):
    """
    Return:
      (years, months, days, full_text, short_text)
    where:
      full_text  -> "24 years 5 months 16 days"
      short_text -> "24 yrs"
    """
    if not dob:
        return None, None, None, None, None

    today = date.today()
    years = today.year - dob.year
    months = today.month - dob.month
    days = today.day - dob.day

    if days < 0:
        months -= 1
        prev_month = today.month - 1 or 12
        prev_year = today.year if today.month > 1 else today.year - 1
        days_in_prev = calendar.monthrange(prev_year, prev_month)[1]
        days += days_in_prev

    if months < 0:
        years -= 1
        months += 12

    parts = []
    if years is not None and years:
        parts.append(f"{years} year{'s' if years != 1 else ''}")
    if months:
        parts.append(f"{months} month{'s' if months != 1 else ''}")
    if days:
        parts.append(f"{days} day{'s' if days != 1 else ''}")

    full_text = " ".join(parts) if parts else None
    short_text = f"{years} yrs" if years is not None else None

    return years, months, days, full_text, short_text


def serialize_patient(p: Patient, db: Session) -> PatientOut:
    years, months, days, full_text, short_text = calc_age(p.dob)
    data = PatientOut.model_validate(p, from_attributes=True)

    # Age fields
    data.age_years = years
    data.age_months = months
    data.age_days = days
    data.age_text = full_text
    data.age_short_text = short_text

    # Resolve doctor & credit master display names
    if p.ref_doctor_id:
        doc = db.query(User).get(p.ref_doctor_id)
        if doc:
            data.ref_doctor_name = doc.name

    if p.credit_payer_id:
        payer = db.query(Payer).get(p.credit_payer_id)
        if payer:
            data.credit_payer_name = payer.name

    if p.credit_tpa_id:
        tpa = db.query(Tpa).get(p.credit_tpa_id)
        if tpa:
            data.credit_tpa_name = tpa.name

    if p.credit_plan_id:
        plan = db.query(CreditPlan).get(p.credit_plan_id)
        if plan:
            data.credit_plan_name = plan.name

    return data


def _safe_str(val: Optional[str]) -> str:
    return val or "—"


def _generate_barcode_data_uri(uhid: str) -> Optional[str]:
    """
    Generate a Code128 barcode (PNG) as data URI using UHID.
    If barcode library is not available, returns None.
    """
    try:
        import barcode  # type: ignore
        from barcode.writer import ImageWriter  # type: ignore
    except Exception:
        return None

    buf = io.BytesIO()
    try:
        Code128 = barcode.get_barcode_class("code128")
        Code128(uhid, writer=ImageWriter()).write(buf)
    except Exception:
        return None

    encoded = base64.b64encode(buf.getvalue()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def _validate_patient_type_in_master(
    db: Session,
    patient_type_value: str,
) -> None:
    """
    Ensure patient_type exists in PatientType master (by code or name) and is active.
    Raises HTTPException 422 if not valid.
    """
    if not patient_type_value:
        raise HTTPException(status_code=422, detail="Patient type is required")

    pt = (db.query(PatientType).filter(
        PatientType.is_active.is_(True),
        (PatientType.code == patient_type_value)
        | (PatientType.name == patient_type_value),
    ).first())
    if not pt:
        raise HTTPException(
            status_code=422,
            detail=
            "Invalid patient type. Please select from Patient Type master.",
        )


def _validate_reference_source_and_doctor(
    db: Session,
    ref_source: Optional[str],
    ref_doctor_id: Optional[int],
) -> Optional[int]:
    """
    When reference source is 'doctor', ensure ref_doctor_id is valid.
    Returns the final ref_doctor_id value (may be None).
    """
    if not ref_source:
        return None

    ref_source_norm = ref_source.strip().lower()
    if ref_source_norm != "doctor":
        # for non-doctor sources, always clear ref_doctor_id
        return None

    if not ref_doctor_id:
        raise HTTPException(
            status_code=422,
            detail=
            "Referring doctor is required when Reference Source is 'Doctor'.",
        )

    doc = db.query(User).get(ref_doctor_id)
    if not doc or not getattr(doc, "is_active", True):
        raise HTTPException(status_code=422,
                            detail="Referring doctor not found")
    # if you have is_doctor flag, you can additionally check that:
    # if not getattr(doc, "is_doctor", False):
    #     raise HTTPException(status_code=422, detail="Selected user is not a doctor")

    return ref_doctor_id


# --------- core patient endpoints ----------


@router.post("", response_model=PatientOut)
@router.post("/", response_model=PatientOut)
def create_patient(
        payload: PatientCreate,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
        request: Request = None,
):
    """
    Create patient.
    Works for both:
      - POST /api/patients
      - POST /api/patients/
    """
    if not has_perm(user, "patients.create"):
        raise HTTPException(status_code=403, detail="Not permitted")

    # Enforce patient_type from master
    _validate_patient_type_in_master(db, payload.patient_type)

    # Validate reference source / doctor combo
    final_ref_doctor_id = _validate_reference_source_and_doctor(
        db,
        payload.ref_source,
        payload.ref_doctor_id,
    )

    # uniqueness checks (phone/email)
    if payload.phone:
        exists = db.query(Patient).filter(
            Patient.phone == payload.phone).first()
        if exists:
            raise HTTPException(status_code=400,
                                detail="Phone already registered")
    if payload.email:
        exists = db.query(Patient).filter(
            Patient.email == payload.email).first()
        if exists:
            raise HTTPException(status_code=400,
                                detail="Email already registered")

    p = Patient(
        uhid="TEMP",
        prefix=(payload.prefix or "").strip(),
        first_name=(payload.first_name or "").strip(),
        last_name=(payload.last_name or "").strip() or None,
        gender=payload.gender,
        dob=payload.dob,
        phone=payload.phone,
        email=payload.email,
        # aadhar_last4=(payload.aadhar_last4 or "")[:4] or None,
        blood_group=payload.blood_group,
        marital_status=payload.marital_status,
        ref_source=payload.ref_source,
        ref_doctor_id=final_ref_doctor_id,
        ref_details=payload.ref_details,
        id_proof_type=payload.id_proof_type,
        id_proof_no=payload.id_proof_no,
        guardian_name=payload.guardian_name,
        guardian_phone=payload.guardian_phone,
        guardian_relation=payload.guardian_relation,
        patient_type=payload.patient_type,
        tag=payload.tag,
        religion=payload.religion,
        occupation=payload.occupation,
        file_number=payload.file_number,
        file_location=payload.file_location,
        credit_type=payload.credit_type,
        credit_payer_id=payload.credit_payer_id,
        credit_tpa_id=payload.credit_tpa_id,
        credit_plan_id=payload.credit_plan_id,
        principal_member_name=payload.principal_member_name,
        principal_member_address=payload.principal_member_address,
        policy_number=payload.policy_number,
        policy_name=payload.policy_name,
        family_id=payload.family_id,
        is_active=True,
    )
    db.add(p)
    db.flush()  # get id
    p.uhid = make_uhid(p.id)

    # create initial address if sent
    if payload.address:
        a = PatientAddress(
            patient_id=p.id,
            type=payload.address.type or "current",
            line1=payload.address.line1,
            line2=payload.address.line2 or "",
            city=payload.address.city or "",
            state=payload.address.state or "",
            pincode=payload.address.pincode or "",
            country=payload.address.country or "India",
        )
        db.add(a)

    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        msg = str(e.orig).lower()
        if "phone" in msg:
            raise HTTPException(status_code=400,
                                detail="Phone already registered")
        if "email" in msg:
            raise HTTPException(status_code=400,
                                detail="Email already registered")
        raise
    db.refresh(p)

    # --- Audit log (CREATE) ---
    meta = get_request_meta(request)
    new_data = instance_to_audit_dict(p)
    log_audit(
        db=db,
        user_id=user.id,
        action="CREATE",
        table_name="patients",
        record_id=p.id,
        old_values=None,
        new_values=new_data,
        ip_address=meta["ip"],
        user_agent=meta["ua"],
    )

    return serialize_patient(p, db)


@router.get("", response_model=List[PatientOut])
@router.get("/", response_model=List[PatientOut])
def list_patients(
        q: Optional[str] = None,
        patient_type: Optional[str] = None,
        limit: int = Query(20, ge=1, le=100),
        offset: int = Query(0, ge=0),
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    """
    List patients with search + filter + pagination.

    - q: search by UHID, name, phone, email
    - patient_type: filter (value from Patient Type master)
    - limit: page size (10 / 20 / 30 / etc up to 100)
    - offset: for pagination (page * limit)
    """
    if not has_perm(user, "patients.view"):
        raise HTTPException(status_code=403, detail="Not permitted")

    qry = db.query(Patient).filter(Patient.is_active.is_(True))
    if q:
        ql = f"%{q.lower()}%"
        qry = qry.filter((Patient.uhid.like(ql))
                         | (Patient.phone.like(ql))
                         | (Patient.email.like(ql))
                         | (Patient.first_name.like(ql))
                         | (Patient.last_name.like(ql)))

    if patient_type:
        qry = qry.filter(Patient.patient_type == patient_type)

    patients = (qry.order_by(
        Patient.id.desc()).offset(offset).limit(limit).all())
    return [serialize_patient(p, db) for p in patients]


# -------- Excel export (report) --------


@router.get("/export", summary="Export patients to Excel")
def export_patients_report(
        from_date: Optional[date] = Query(None, alias="from_date"),
        to_date: Optional[date] = Query(None, alias="to_date"),
        patient_type: Optional[str] = None,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    """
    Generate Excel report based on:
      - from_date, to_date (by created_at)
      - patient_type (optional)

    If dates are not provided, defaults to last 30 days.
    """
    if not has_perm(user, "patients.view"):
        raise HTTPException(status_code=403, detail="Not permitted")

    # ---- Default date range / validation ----
    today = date.today()
    if to_date is None:
        to_date = today

    if from_date is None:
        # default: last 30 days window
        from_date = to_date - timedelta(days=30)

    if to_date < from_date:
        raise HTTPException(
            status_code=422,
            detail="To Date must be on or after From Date",
        )

    start_dt = datetime.combine(from_date, datetime.min.time())
    end_dt = datetime.combine(to_date + timedelta(days=1), datetime.min.time())

    try:
        # ---- Query patients ----
        qry = db.query(Patient).filter(
            Patient.is_active.is_(True),
            Patient.created_at >= start_dt,
            Patient.created_at < end_dt,
        )

        if patient_type:
            qry = qry.filter(Patient.patient_type == patient_type)

        patients = qry.order_by(Patient.id.asc()).all()

        # ---- Build workbook ----
        wb = Workbook()
        ws = wb.active
        ws.title = "Patients"

        headers = [
            "UHID",
            "Prefix",
            "First Name",
            "Last Name",
            "Gender",
            "DOB",
            "Age",
            "Marital Status",
            "Mobile",
            "Email",
            "Patient Type",
            "Created At",
        ]
        ws.append(headers)

        for p in patients:
            _, _, _, age_text, _ = calc_age(p.dob)
            ws.append([
                p.uhid,
                p.prefix or "",
                p.first_name or "",
                p.last_name or "",
                p.gender or "",
                p.dob.isoformat() if p.dob else "",
                age_text or "",
                p.marital_status or "",
                p.phone or "",
                p.email or "",
                p.patient_type or "",
                p.created_at.isoformat()
                if getattr(p, "created_at", None) else "",
            ])

        # simple column width auto-fit
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in ws[col_letter])
            ws.column_dimensions[col_letter].width = max(max_len + 2, 12)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

    except Exception as e:
        # IMPORTANT: this will bubble proper error msg to frontend
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Excel export failed: {e}",
        )

    filename = f"patients_{from_date.isoformat()}_to_{to_date.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename=\"{filename}\"'
        },
    )


@router.get("/{patient_id}", response_model=PatientOut)
def get_patient(
        patient_id: int,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "patients.view"):
        raise HTTPException(status_code=403, detail="Not permitted")
    p = db.query(Patient).get(patient_id)
    if not p or not p.is_active:
        raise HTTPException(status_code=404, detail="Not found")
    return serialize_patient(p, db)


@router.put("/{patient_id}", response_model=PatientOut)
def update_patient(
        patient_id: int,
        payload: PatientUpdate,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
        request: Request = None,
):
    if not has_perm(user, "patients.update"):
        raise HTTPException(status_code=403, detail="Not permitted")

    p = db.query(Patient).get(patient_id)
    if not p or not p.is_active:
        raise HTTPException(status_code=404, detail="Not found")

    old_data = instance_to_audit_dict(p)

    data = payload.dict(exclude_unset=True)

    # uniqueness when changing phone/email
    if "phone" in data:
        new_phone = data["phone"]
        if new_phone:
            exists = (db.query(Patient).filter(
                Patient.phone == new_phone,
                Patient.id != patient_id,
            ).first())
            if exists:
                raise HTTPException(status_code=400,
                                    detail="Phone already registered")

    if "email" in data:
        new_email = data["email"]
        if new_email:
            exists = (db.query(Patient).filter(
                Patient.email == new_email,
                Patient.id != patient_id,
            ).first())
            if exists:
                raise HTTPException(status_code=400,
                                    detail="Email already registered")

    # validate patient_type if changed
    if "patient_type" in data and data["patient_type"] is not None:
        _validate_patient_type_in_master(db, data["patient_type"])

    # reference source / doctor validation using final values
    ref_source_val = data.get("ref_source", p.ref_source)
    ref_doctor_val = data.get("ref_doctor_id", p.ref_doctor_id)
    final_ref_doctor_id = _validate_reference_source_and_doctor(
        db,
        ref_source_val,
        ref_doctor_val,
    )
    if ref_source_val and ref_source_val.strip().lower() != "doctor":
        data["ref_doctor_id"] = None
    else:
        data["ref_doctor_id"] = final_ref_doctor_id

    # normalize some text fields
    if "prefix" in data and isinstance(data["prefix"], str):
        data["prefix"] = data["prefix"].strip()
    if "first_name" in data and isinstance(data["first_name"], str):
        data["first_name"] = data["first_name"].strip()
    if "last_name" in data and isinstance(data["last_name"], str):
        data["last_name"] = data["last_name"].strip()

    for k, v in data.items():
        setattr(p, k, v)

    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        msg = str(e.orig).lower()
        if "phone" in msg:
            raise HTTPException(status_code=400,
                                detail="Phone already registered")
        if "email" in msg:
            raise HTTPException(status_code=400,
                                detail="Email already registered")
        raise
    db.refresh(p)

    # --- Audit log (UPDATE) ---
    meta = get_request_meta(request)
    new_data = instance_to_audit_dict(p)
    log_audit(
        db=db,
        user_id=user.id,
        action="UPDATE",
        table_name="patients",
        record_id=p.id,
        old_values=old_data,
        new_values=new_data,
        ip_address=meta["ip"],
        user_agent=meta["ua"],
    )

    return serialize_patient(p, db)


@router.patch("/{patient_id}/deactivate")
def deactivate_patient(
        patient_id: int,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
        request: Request = None,
):
    if not has_perm(user, "patients.deactivate"):
        raise HTTPException(status_code=403, detail="Not permitted")
    p = db.query(Patient).get(patient_id)
    if not p:
        raise HTTPException(status_code=404, detail="Not found")

    old_data = instance_to_audit_dict(p)

    p.is_active = False
    db.commit()
    db.refresh(p)

    # --- Audit log (DELETE / soft) ---
    meta = get_request_meta(request)
    new_data = instance_to_audit_dict(p)
    log_audit(
        db=db,
        user_id=user.id,
        action="DELETE",
        table_name="patients",
        record_id=p.id,
        old_values=old_data,
        new_values=new_data,
        ip_address=meta["ip"],
        user_agent=meta["ua"],
    )

    return {"message": "Deactivated"}


# -------- addresses --------


@router.post("/{patient_id}/addresses", response_model=AddressOut)
def add_address(
        patient_id: int,
        payload: AddressIn,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
        request: Request = None,
):
    if not has_perm(user, "patients.addresses.create"):
        raise HTTPException(status_code=403, detail="Not permitted")
    p = db.query(Patient).get(patient_id)
    if not p:
        raise HTTPException(status_code=404, detail="Not found")
    a = PatientAddress(
        patient_id=patient_id,
        type=payload.type or "current",
        line1=payload.line1,
        line2=payload.line2,
        city=payload.city,
        state=payload.state,
        pincode=payload.pincode,
        country=payload.country or "India",
    )
    db.add(a)
    db.commit()
    db.refresh(a)

    # --- Audit log (CREATE address) ---
    meta = get_request_meta(request)
    new_data = instance_to_audit_dict(a)
    log_audit(
        db=db,
        user_id=user.id,
        action="CREATE",
        table_name="patient_addresses",
        record_id=a.id,
        old_values=None,
        new_values=new_data,
        ip_address=meta["ip"],
        user_agent=meta["ua"],
    )

    return a


@router.get("/{patient_id}/addresses", response_model=List[AddressOut])
def list_addresses(
        patient_id: int,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "patients.addresses.view"):
        raise HTTPException(status_code=403, detail="Not permitted")
    return (db.query(PatientAddress).filter(
        PatientAddress.patient_id == patient_id).order_by(
            PatientAddress.id.desc()).all())


@router.put("/addresses/{addr_id}", response_model=AddressOut)
def update_address(
        addr_id: int,
        payload: AddressIn,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
        request: Request = None,
):
    if not has_perm(user, "patients.addresses.update"):
        raise HTTPException(status_code=403, detail="Not permitted")
    a = db.query(PatientAddress).get(addr_id)
    if not a:
        raise HTTPException(status_code=404, detail="Address not found")

    old_data = instance_to_audit_dict(a)

    for k, v in payload.dict(exclude_unset=True).items():
        setattr(a, k, v)
    db.commit()
    db.refresh(a)

    # --- Audit log (UPDATE address) ---
    meta = get_request_meta(request)
    new_data = instance_to_audit_dict(a)
    log_audit(
        db=db,
        user_id=user.id,
        action="UPDATE",
        table_name="patient_addresses",
        record_id=a.id,
        old_values=old_data,
        new_values=new_data,
        ip_address=meta["ip"],
        user_agent=meta["ua"],
    )

    return a


@router.delete("/addresses/{addr_id}")
def delete_address(
        addr_id: int,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
        request: Request = None,
):
    if not has_perm(user, "patients.addresses.delete"):
        raise HTTPException(status_code=403, detail="Not permitted")
    a = db.query(PatientAddress).get(addr_id)
    if not a:
        raise HTTPException(status_code=404, detail="Address not found")

    old_data = instance_to_audit_dict(a)

    db.delete(a)
    db.commit()

    # --- Audit log (DELETE address) ---
    meta = get_request_meta(request)
    log_audit(
        db=db,
        user_id=user.id,
        action="DELETE",
        table_name="patient_addresses",
        record_id=addr_id,
        old_values=old_data,
        new_values=None,
        ip_address=meta["ip"],
        user_agent=meta["ua"],
    )

    return {"message": "Deleted"}


# -------- documents (upload + list + file) --------


@router.post("/{patient_id}/documents", response_model=DocumentOut)
async def upload_document(
        patient_id: int,
        type: str = Form("other"),
        file: UploadFile = File(...),
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
        request: Request = None,
):
    if not has_perm(user, "patients.attachments.manage"):
        raise HTTPException(status_code=403, detail="Not permitted")
    p = db.query(Patient).get(patient_id)
    if not p:
        raise HTTPException(status_code=404, detail="Not found")

    ext = os.path.splitext(file.filename or "")[1]
    safe_name = f"p{patient_id}_{os.urandom(6).hex()}{ext}"
    dest = UPLOAD_DIR.joinpath(safe_name)
    content = await file.read()
    dest.write_bytes(content)

    doc = PatientDocument(
        patient_id=patient_id,
        type=type,
        filename=file.filename or safe_name,
        mime=file.content_type or "",
        size=len(content),
        storage_path=str(dest),
        uploaded_by=user.id,
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)

    # --- Audit log (CREATE document) ---
    meta = get_request_meta(request)
    new_data = instance_to_audit_dict(doc)
    log_audit(
        db=db,
        user_id=user.id,
        action="CREATE",
        table_name="patient_documents",
        record_id=doc.id,
        old_values=None,
        new_values=new_data,
        ip_address=meta["ip"],
        user_agent=meta["ua"],
    )

    return doc


@router.get("/{patient_id}/documents", response_model=List[DocumentOut])
def list_documents(
        patient_id: int,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "patients.view"):
        raise HTTPException(status_code=403, detail="Not permitted")
    return (db.query(PatientDocument).filter(
        PatientDocument.patient_id == patient_id).order_by(
            PatientDocument.id.desc()).all())


@router.get("/documents/{doc_id}/file")
def get_document_file(
        doc_id: int,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    """
    Stream patient document file for preview / download.
    URL: GET /api/patients/documents/{doc_id}/file
    """
    if not has_perm(user, "patients.view"):
        raise HTTPException(status_code=403, detail="Not permitted")

    doc = db.query(PatientDocument).get(doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    path = Path(doc.storage_path)
    if not path.exists():
        raise HTTPException(status_code=404, detail="File not found on server")

    media_type = doc.mime or "application/octet-stream"
    headers = {
        "Content-Disposition": f'inline; filename="{doc.filename}"',
    }

    return StreamingResponse(
        path.open("rb"),
        media_type=media_type,
        headers=headers,
    )


# -------- consents --------


@router.post("/{patient_id}/consents", response_model=ConsentOut)
def create_consent(
        patient_id: int,
        payload: ConsentIn = Body(...),
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
        request: Request = None,
):
    if not has_perm(user, "patients.consents.create"):
        raise HTTPException(status_code=403, detail="Not permitted")

    p = db.query(Patient).get(patient_id)
    if not p:
        raise HTTPException(status_code=404, detail="Patient not found")

    c = PatientConsent(
        patient_id=patient_id,
        type=payload.type,
        text=payload.text,
    )
    db.add(c)
    db.commit()
    db.refresh(c)

    # --- Audit log (CREATE consent) ---
    meta = get_request_meta(request)
    new_data = instance_to_audit_dict(c)
    log_audit(
        db=db,
        user_id=user.id,
        action="CREATE",
        table_name="patient_consents",
        record_id=c.id,
        old_values=None,
        new_values=new_data,
        ip_address=meta["ip"],
        user_agent=meta["ua"],
    )

    return c


@router.get("/{patient_id}/consents", response_model=List[ConsentOut])
def list_consents(
        patient_id: int,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    if not has_perm(user, "patients.consents.view"):
        raise HTTPException(status_code=403, detail="Not permitted")

    return (db.query(PatientConsent).filter(
        PatientConsent.patient_id == patient_id).order_by(
            PatientConsent.id.desc()).all())


# -------- Patient Slip PDF (with barcode) --------


@router.get("/{patient_id}/slip")
def print_patient_slip(
        patient_id: int,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    """
    Patient slip PDF (or HTML fallback) including:
      - Barcode (UHID)
      - Prefix + Patient name
      - Patient type
      - Mobile, Gender, Email
    Frontend can call this immediately after registration.
    """
    if not has_perm(user, "patients.view"):
        raise HTTPException(status_code=403, detail="Not permitted")

    p = db.query(Patient).get(patient_id)
    if not p or not p.is_active:
        raise HTTPException(status_code=404, detail="Not found")

    barcode_data_uri = _generate_barcode_data_uri(p.uhid)

    full_name = " ".join(part for part in [
        p.prefix or "",
        p.first_name or "",
        p.last_name or "",
    ] if part)

    html = f"""
<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8" />
  <title>Patient Slip - {p.uhid}</title>
  <style>
    body {{
      font-family: system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      font-size: 11px;
      margin: 12px;
      color: #111827;
    }}
    .slip {{
      border-radius: 8px;
      border: 1px solid #e5e7eb;
      padding: 10px 12px;
      width: 320px;
    }}
    .header {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      margin-bottom: 6px;
    }}
    .uhid {{
      font-weight: 600;
      font-size: 12px;
    }}
    .label {{
      font-weight: 500;
    }}
    .value {{
      color: #111827;
    }}
    .row {{
      display: flex;
      justify-content: space-between;
      margin-bottom: 2px;
    }}
    .barcode {{
      margin-top: 4px;
      text-align: center;
    }}
    .barcode img {{
      max-width: 100%;
      height: 40px;
    }}
  </style>
</head>
<body>
  <div class="slip">
    <div class="header">
      <div>
        <div class="uhid">UHID: {p.uhid}</div>
        <div style="font-size:10px;color:#6b7280;">Patient Slip</div>
      </div>
      <div style="font-size:9px;color:#6b7280;">Date: {(p.created_at or datetime.utcnow()).strftime("%d-%m-%Y")}</div>
    </div>

    <div class="row">
      <div><span class="label">Name:</span> <span class="value">{full_name}</span></div>
    </div>
    <div class="row">
      <div><span class="label">Gender:</span> <span class="value">{_safe_str(p.gender)}</span></div>
      <div><span class="label">Type:</span> <span class="value">{_safe_str(p.patient_type)}</span></div>
    </div>
    <div class="row">
      <div><span class="label">Mobile:</span> <span class="value">{_safe_str(p.phone)}</span></div>
    </div>
    <div class="row">
      <div><span class="label">Email:</span> <span class="value">{_safe_str(p.email)}</span></div>
    </div>

    <div class="barcode">
      {"<img src='" + barcode_data_uri + "' alt='UHID Barcode' />" if barcode_data_uri else "<div class='value'>UHID: " + p.uhid + "</div>"}
    </div>
  </div>
</body>
</html>
    """.strip()

    # Lazy + safe import of WeasyPrint
    try:
        from weasyprint import HTML as _HTML  # type: ignore

        HTML = _HTML
    except Exception:
        HTML = None

    if HTML is None:
        # Fallback: return HTML so browser print dialog can be used
        return StreamingResponse(
            io.BytesIO(html.encode("utf-8")),
            media_type="text/html; charset=utf-8",
        )

    pdf_bytes = HTML(string=html).write_pdf()
    filename = f"patient-slip-{p.uhid}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


# -------- existing Patient Info print --------


@router.get("/{patient_id}/print-info")
def print_patient_info(
        patient_id: int,
        db: Session = Depends(get_db),
        user: User = Depends(auth_current_user),
):
    """
    Print patient info:
    - If WeasyPrint and OS libs available -> return PDF
    - Else -> return HTML (so browser print works)
    """
    if not has_perm(user, "patients.view"):
        raise HTTPException(status_code=403, detail="Not permitted")

    p = db.query(Patient).get(patient_id)
    if not p or not p.is_active:
        raise HTTPException(status_code=404, detail="Not found")

    _, _, _, age_text, _ = calc_age(p.dob)
    addr = p.addresses[0] if p.addresses else None

    def safe(val: Optional[str]) -> str:
        return val or "—"

    address_block = "—"
    if addr:
        parts = [
            addr.line1 or "",
            addr.line2 or "",
            " ".join(x for x in [addr.city, addr.state, addr.pincode]
                     if x and x.strip()),
            addr.country or "",
        ]
        address_block = "<br/>".join([p for p in parts if p])

    html = f"""
<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8" />
  <title>Patient Info - {p.uhid}</title>
  <style>
    body {{
      font-family: system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      font-size: 12px;
      margin: 16px;
      color: #111827;
    }}
    .card {{
      border-radius: 8px;
      border: 1px solid #e5e7eb;
      padding: 12px 16px;
      margin-bottom: 8px;
    }}
    .title {{
      font-size: 16px;
      font-weight: 600;
      margin-bottom: 4px;
    }}
    .muted {{
      color: #6b7280;
      font-size: 11px;
    }}
    .grid {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 4px 12px;
      margin-top: 6px;
    }}
    .label {{
      font-weight: 500;
    }}
    .value {{
      color: #111827;
    }}
  </style>
</head>
<body>
  <div class="card">
    <div class="title">Patient Information</div>
    <div class="muted">UHID: {p.uhid}</div>

    <div class="grid">
      <div><span class="label">Name:</span> <span class="value">{safe(p.prefix)} {safe(p.first_name)} {p.last_name or ""}</span></div>
      <div><span class="label">Gender:</span> <span class="value">{safe(p.gender)}</span></div>
      <div><span class="label">DOB:</span> <span class="value">{p.dob or "—"}</span></div>
      <div><span class="label">Age:</span> <span class="value">{age_text or "—"}</span></div>
      <div><span class="label">Blood Group:</span> <span class="value">{safe(p.blood_group)}</span></div>
      <div><span class="label">Marital Status:</span> <span class="value">{safe(p.marital_status)}</span></div>
      <div><span class="label">Mobile:</span> <span class="value">{safe(p.phone)}</span></div>
      <div><span class="label">Email:</span> <span class="value">{safe(p.email)}</span></div>
      <div><span class="label">Patient Type:</span> <span class="value">{safe(p.patient_type)}</span></div>
      <div><span class="label">Tag:</span> <span class="value">{safe(p.tag)}</span></div>
      <div><span class="label">Religion:</span> <span class="value">{safe(p.religion)}</span></div>
      <div><span class="label">Occupation:</span> <span class="value">{safe(p.occupation)}</span></div>
      <div><span class="label">File No:</span> <span class="value">{safe(p.file_number)}</span></div>
      <div><span class="label">File Location:</span> <span class="value">{safe(p.file_location)}</span></div>
      <div><span class="label">Guardian:</span> <span class="value">{safe(p.guardian_name)}</span></div>
      <div><span class="label">Guardian Phone:</span> <span class="value">{safe(p.guardian_phone)}</span></div>
      <div><span class="label">ID Proof:</span> <span class="value">{safe(p.id_proof_type)} {safe(p.id_proof_no)}</span></div>
    
    </div>
  </div>

  <div class="card">
    <div class="title">Address</div>
    <div class="muted">Primary address</div>
    <div class="value">
      {address_block}
    </div>
  </div>
</body>
</html>
    """.strip()

    # Lazy + safe import of WeasyPrint
    try:
        from weasyprint import HTML as _HTML  # type: ignore

        HTML = _HTML
    except Exception:
        HTML = None

    if HTML is None:
        # Fallback: return HTML so browser print dialog can be used
        return StreamingResponse(
            io.BytesIO(html.encode("utf-8")),
            media_type="text/html; charset=utf-8",
        )

    pdf_bytes = HTML(string=html).write_pdf()
    filename = f"patient-{p.uhid}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )
