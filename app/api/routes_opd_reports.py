from __future__ import annotations

from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, selectinload

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.api.deps import get_db, current_user
from app.models.user import User
from app.models.department import Department
from app.models.patient import Patient
from app.models.opd import (
    OpdSchedule,
    Appointment,
    Visit,
    Vitals,
    FollowUp,
    Prescription,
    PrescriptionItem,
    LabOrder,
    RadiologyOrder,
)

IST = ZoneInfo("Asia/Kolkata")

router = APIRouter(prefix="/opd/reports", tags=["OPD Reports"])


# ----------------------------
# permissions (safe helper)
# ----------------------------
def _parse_perms(v):
    if v is None:
        return []

    # list of strings OR list of dicts
    if isinstance(v, list) or isinstance(v, tuple) or isinstance(v, set):
        out = []
        for x in v:
            if isinstance(x, str):
                out.append(x)
            elif isinstance(x, dict):
                code = x.get("code") or x.get("perm") or x.get("name")
                if code:
                    out.append(str(code))
            else:
                out.append(str(x))
        return [p.strip() for p in out if str(p).strip()]

    # comma-separated string
    if isinstance(v, str):
        return [x.strip() for x in v.split(",") if x.strip()]

    # dict
    if isinstance(v, dict):
        # { permissions: [...] } or { codes: [...] }
        for k in ("permissions", "codes", "perms"):
            if k in v:
                return _parse_perms(v.get(k))
        code = v.get("code") or v.get("perm") or v.get("name")
        return [str(code).strip()] if code else []

    return [str(v).strip()]



def _need_any(user: User, codes: list[str]) -> None:
    if getattr(user, "is_admin", False):
        return
    have = {p.code for r in (user.roles or []) for p in (getattr(r, "permissions", []) or [])}
    if have.intersection(set(codes)):
        return
    raise HTTPException(status_code=403, detail="Not permitted")


# ----------------------------
# formatting helpers
# ----------------------------
def _s(v: Any, dash: str = "—") -> str:
    if v is None:
        return dash
    s = str(v).strip()
    return s if s else dash


def _dt_ist(dt):
    """
    Convert dt to IST and return NAIVE datetime (tzinfo=None)
    because Excel/openpyxl does not support tz-aware datetimes.
    """
    if not dt:
        return None

    # if stored naive, assume UTC
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    ist = dt.astimezone(IST)

    # ✅ Excel needs tzinfo=None
    return ist.replace(tzinfo=None)

def _excel_safe(v):
    if isinstance(v, datetime):
        # ensure tzinfo is removed
        return v.replace(tzinfo=None) if v.tzinfo else v
    return v



def _time_str(t: Any) -> str:
    if not t:
        return "—"
    if isinstance(t, time):
        return t.strftime("%H:%M")
    try:
        return str(t)
    except Exception:
        return "—"


def _patient_uid(p: Optional[Patient]) -> str:
    if not p:
        return "—"
    # try common fields without breaking your build
    for k in ("uhid", "patient_no", "patient_code", "mrn", "uid", "unique_no"):
        val = getattr(p, k, None)
        if val:
            return str(val)
    return str(getattr(p, "id", "—"))


def _patient_name(p: Optional[Patient]) -> str:
    if not p:
        return "—"
    for k in ("full_name", "name"):
        val = getattr(p, k, None)
        if val:
            return str(val)
    fn = getattr(p, "first_name", "") or ""
    ln = getattr(p, "last_name", "") or ""
    nm = (fn + " " + ln).strip()
    return nm if nm else "—"


def _doctor_name(u: Any) -> str:
    if not u:
        return "—"
    for k in ("full_name", "name", "display_name"):
        val = getattr(u, k, None)
        if val:
            return str(val)
    fn = getattr(u, "first_name", "") or ""
    ln = getattr(u, "last_name", "") or ""
    nm = (fn + " " + ln).strip()
    return nm if nm else _s(getattr(u, "username", None))


def _dept_name(d: Any) -> str:
    if not d:
        return "—"
    return _s(getattr(d, "name", None))


def _calc_age(p: Optional[Patient]) -> str:
    if not p:
        return "—"
    dob = getattr(p, "dob", None) or getattr(p, "date_of_birth", None)
    if not dob:
        return "—"
    try:
        if isinstance(dob, datetime):
            dob = dob.date()
        today = date.today()
        years = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        return str(max(years, 0))
    except Exception:
        return "—"


def _truncate(v: Any, limit: int = 32000) -> str:
    s = "" if v is None else str(v)
    s = s.strip()
    if not s:
        return "—"
    return s[:limit]


# ----------------------------
# excel helpers
# ----------------------------
HEADER_FILL = PatternFill("solid", fgColor="111827")  # slate-900
HEADER_FONT = Font(bold=True, color="FFFFFF")
ROW_FONT = Font(color="111827")
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_sheet(ws, title: str, headers: List[str], rows: List[List[Any]]) -> None:
    ws.title = title

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    # ✅ append rows once (Excel-safe)
    for r in rows:
        ws.append([_excel_safe(x) for x in r])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # style body + widths
    col_max: Dict[int, int] = {i: len(h) for i, h in enumerate(headers, start=1)}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for c in row:
            c.font = ROW_FONT
            c.alignment = WRAP
            try:
                val_len = len(str(c.value)) if c.value is not None else 0
            except Exception:
                val_len = 0
            col_max[c.column] = max(col_max.get(c.column, 10), min(val_len, 60))

    for col_idx, mx in col_max.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(mx + 2, 60))


@router.get("/opd.xlsx")
def export_opd_excel(
    date_from: date = Query(...),
    date_to: date = Query(...),
    doctor_user_id: Optional[int] = Query(None),
    department_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None, description="Appointment status filter (booked/checked_in/etc)"),
    include_notes: bool = Query(True),
    db: Session = Depends(get_db),
    user: User = Depends(current_user),
):
    # ✅ permissions: adjust to your project codes if needed
    _need_any(
        user,
        [
            "opd.reports.view",
            "opd.view",
            "opd.visits.view",
            "opd.appointments.view",
            "visits.view",
            "appointments.view",
        ],
    )

    if date_to < date_from:
        raise HTTPException(status_code=400, detail="date_to must be >= date_from")

    dt_from = datetime.combine(date_from, time.min).replace(tzinfo=timezone.utc)
    dt_to = datetime.combine(date_to + timedelta(days=1), time.min).replace(tzinfo=timezone.utc)

    # ----------------------------
    # VISITS (main)
    # ----------------------------
    vq = (
        db.query(Visit)
        .options(
            selectinload(Visit.patient),
            selectinload(Visit.doctor),
            selectinload(Visit.department),
            selectinload(Visit.appointment),
        )
        .filter(Visit.visit_at >= dt_from, Visit.visit_at < dt_to)
    )
    if doctor_user_id:
        vq = vq.filter(Visit.doctor_user_id == doctor_user_id)
    if department_id:
        vq = vq.filter(Visit.department_id == department_id)

    visits: List[Visit] = vq.order_by(Visit.visit_at.asc()).all()
    visit_ids = [v.id for v in visits]
    appt_ids = [v.appointment_id for v in visits if v.appointment_id]

    # ----------------------------
    # VITALS (latest per appointment)
    # ----------------------------
    vitals_map: Dict[int, Vitals] = {}
    if appt_ids:
        vs = (
            db.query(Vitals)
            .filter(Vitals.appointment_id.in_(appt_ids))
            .order_by(Vitals.created_at.asc())
            .all()
        )
        for x in vs:
            if x.appointment_id:
                vitals_map[x.appointment_id] = x  # keep latest

    # ----------------------------
    # RX + ITEMS
    # ----------------------------
    rx_map: Dict[int, Prescription] = {}
    rx_items_by_visit: Dict[int, List[PrescriptionItem]] = {}
    if visit_ids:
        rxs = (
            db.query(Prescription)
            .options(selectinload(Prescription.items), selectinload(Prescription.signer))
            .filter(Prescription.visit_id.in_(visit_ids))
            .all()
        )
        for rx in rxs:
            rx_map[rx.visit_id] = rx
            rx_items_by_visit[rx.visit_id] = list(rx.items or [])

    def _rx_summary(visit_id: int) -> str:
        items = rx_items_by_visit.get(visit_id) or []
        if not items:
            return "—"
        parts = []
        for it in items:
            parts.append(
                f"{_s(getattr(it,'drug_name',None))}"
                f"{(' ' + _s(getattr(it,'strength',None))) if getattr(it,'strength',None) else ''}"
                f" | {_s(getattr(it,'frequency',None))}"
                f" | {int(getattr(it,'duration_days',0) or 0)}d"
                f" | Qty:{int(getattr(it,'quantity',0) or 0)}"
            )
        s = "; ".join(parts)
        return _truncate(s, 32000)

    # ----------------------------
    # LAB / RIS
    # ----------------------------
    lab_by_visit: Dict[int, List[str]] = {}
    ris_by_visit: Dict[int, List[str]] = {}

    if visit_ids:
        lab_orders = (
            db.query(LabOrder)
            .options(selectinload(LabOrder.test))
            .filter(LabOrder.visit_id.in_(visit_ids))
            .order_by(LabOrder.ordered_at.asc())
            .all()
        )
        for o in lab_orders:
            name = getattr(getattr(o, "test", None), "name", None) or getattr(getattr(o, "test", None), "code", None)
            lab_by_visit.setdefault(o.visit_id, []).append(f"{_s(name)} ({_s(o.status)})")

        ris_orders = (
            db.query(RadiologyOrder)
            .options(selectinload(RadiologyOrder.test))
            .filter(RadiologyOrder.visit_id.in_(visit_ids))
            .order_by(RadiologyOrder.ordered_at.asc())
            .all()
        )
        for o in ris_orders:
            t = getattr(o, "test", None)
            name = getattr(t, "name", None) or getattr(t, "code", None)
            mod = getattr(t, "modality", None)
            label = f"{_s(name)}"
            if mod:
                label += f" [{mod}]"
            ris_by_visit.setdefault(o.visit_id, []).append(f"{label} ({_s(o.status)})")

    # ----------------------------
    # FOLLOWUPS
    # ----------------------------
    followups: List[FollowUp] = []
    followup_by_visit: Dict[int, FollowUp] = {}
    if visit_ids:
        followups = (
            db.query(FollowUp)
            .options(selectinload(FollowUp.appointment))
            .filter(FollowUp.source_visit_id.in_(visit_ids))
            .order_by(FollowUp.created_at.asc())
            .all()
        )
        for fu in followups:
            followup_by_visit[fu.source_visit_id] = fu  # if multiple, keep last

    # ----------------------------
    # APPOINTMENTS (sheet)
    # ----------------------------
    aq = (
        db.query(Appointment)
        .options(
            selectinload(Appointment.patient),
            selectinload(Appointment.doctor),
            selectinload(Appointment.department),
        )
        .filter(Appointment.date >= date_from, Appointment.date <= date_to)
    )
    if doctor_user_id:
        aq = aq.filter(Appointment.doctor_user_id == doctor_user_id)
    if department_id:
        aq = aq.filter(Appointment.department_id == department_id)
    if status:
        aq = aq.filter(Appointment.status == status)

    appts: List[Appointment] = aq.order_by(Appointment.date.asc(), Appointment.queue_no.asc()).all()

    # ----------------------------
    # SCHEDULES (sheet)
    # ----------------------------
    sq = (
        db.query(OpdSchedule)
        .options(selectinload(OpdSchedule.doctor))
        .filter(OpdSchedule.is_active == True)  # noqa: E712
    )
    if doctor_user_id:
        sq = sq.filter(OpdSchedule.doctor_user_id == doctor_user_id)
    schedules: List[OpdSchedule] = sq.order_by(OpdSchedule.doctor_user_id.asc(), OpdSchedule.weekday.asc()).all()

    # ----------------------------
    # Build Workbook
    # ----------------------------
    wb = Workbook()
    wb.remove(wb.active)

    # 1) OPD Register (Visits)
    visit_headers = [
        "Visit DateTime (IST)",
        "OP No / Episode",
        "Appointment Date",
        "Appointment Type",
        "Queue No",
        "Slot Start",
        "Slot End",
        "Appointment Status",
        "Department",
        "Doctor",
        "Patient UHID",
        "Patient Name",
        "Age",
        "Sex",
        "Phone",
        "Chief Complaint",
        "Symptoms",
        "Provisional Dx",
        "Final Dx",
        "Advice",
        "Followup Plan",
        "Vitals Summary",
        "Prescription Summary",
        "Lab Orders",
        "Radiology Orders",
        "Followup Due Date",
        "Followup Status",
        "Followup Note",
    ]

    visit_rows: List[List[Any]] = []
    for v in visits:
        p = getattr(v, "patient", None)
        d = getattr(v, "doctor", None)
        dep = getattr(v, "department", None)
        ap = getattr(v, "appointment", None)

        vit = vitals_map.get(getattr(v, "appointment_id", None) or -1)
        vit_s = "—"
        if vit:
            vit_s = (
                f"HT:{_s(getattr(vit,'height_cm',None))}cm, "
                f"WT:{_s(getattr(vit,'weight_kg',None))}kg, "
                f"BP:{_s(getattr(vit,'bp_systolic',None))}/{_s(getattr(vit,'bp_diastolic',None))}, "
                f"P:{_s(getattr(vit,'pulse',None))}, "
                f"SpO2:{_s(getattr(vit,'spo2',None))}, "
                f"T:{_s(getattr(vit,'temp_c',None))}"
            )

        fu = followup_by_visit.get(v.id)

        if include_notes:
            cc = _truncate(getattr(v, "chief_complaint", None))
            sym = _truncate(getattr(v, "symptoms", None))
            pdx = _truncate(getattr(v, "provisional_diagnosis", None))
            fdx = _truncate(getattr(v, "final_diagnosis", None))
            advice = _truncate(getattr(v, "advice", None))
            fplan = _truncate(getattr(v, "followup_plan", None))
        else:
            cc = sym = pdx = fdx = advice = fplan = "—"

        visit_rows.append([
            _dt_ist(getattr(v, "visit_at", None)),
            _s(getattr(v, "episode_id", None) or getattr(v, "op_no", None)),
            _s(getattr(ap, "date", None)),
            _s(getattr(ap, "appointment_type", None)),
            _s(getattr(ap, "queue_no", None)),
            _time_str(getattr(ap, "slot_start", None)),
            _time_str(getattr(ap, "slot_end", None)),
            _s(getattr(ap, "status", None)),
            _dept_name(dep),
            _doctor_name(d),
            _patient_uid(p),
            _patient_name(p),
            _calc_age(p),
            _s(getattr(p, "sex", None) or getattr(p, "gender", None)),
            _s(getattr(p, "phone", None) or getattr(p, "mobile", None)),
            cc,
            sym,
            pdx,
            fdx,
            advice,
            fplan,
            _truncate(vit_s),
            _rx_summary(v.id),
            _truncate("; ".join(lab_by_visit.get(v.id, []) or []) or "—"),
            _truncate("; ".join(ris_by_visit.get(v.id, []) or []) or "—"),
            _s(getattr(fu, "due_date", None)),
            _s(getattr(fu, "status", None)),
            _truncate(getattr(fu, "note", None)) if include_notes else "—",
        ])

    ws_visits = wb.create_sheet()
    _write_sheet(ws_visits, "OPD Register", visit_headers, visit_rows)

    # 2) Appointments
    appt_headers = [
        "Date",
        "Appointment Type",
        "Queue No",
        "Slot Start",
        "Slot End",
        "Status",
        "Purpose",
        "Department",
        "Doctor",
        "Patient UHID",
        "Patient Name",
        "Phone",
        "Created At (IST)",
    ]
    appt_rows: List[List[Any]] = []
    for a in appts:
        appt_rows.append([
            _s(getattr(a, "date", None)),
            _s(getattr(a, "appointment_type", None)),
            _s(getattr(a, "queue_no", None)),
            _time_str(getattr(a, "slot_start", None)),
            _time_str(getattr(a, "slot_end", None)),
            _s(getattr(a, "status", None)),
            _truncate(getattr(a, "purpose", None), 500),
            _dept_name(getattr(a, "department", None)),
            _doctor_name(getattr(a, "doctor", None)),
            _patient_uid(getattr(a, "patient", None)),
            _patient_name(getattr(a, "patient", None)),
            _s(getattr(getattr(a, "patient", None), "phone", None) or getattr(getattr(a, "patient", None), "mobile", None)),
            _dt_ist(getattr(a, "created_at", None)),
        ])
    ws_appts = wb.create_sheet()
    _write_sheet(ws_appts, "Appointments", appt_headers, appt_rows)

    # 3) Vitals (from the vitals rows we already pulled)
    vit_headers = [
        "Appointment ID",
        "Patient UHID",
        "Patient Name",
        "Height (cm)",
        "Weight (kg)",
        "BP Sys",
        "BP Dia",
        "Pulse",
        "RR",
        "Temp (C)",
        "SpO2",
        "Notes",
        "Created At (IST)",
    ]
    vit_rows: List[List[Any]] = []
    for appt_id, v in vitals_map.items():
        p = getattr(v, "patient", None)
        vit_rows.append([
            appt_id,
            _patient_uid(p),
            _patient_name(p),
            _s(getattr(v, "height_cm", None)),
            _s(getattr(v, "weight_kg", None)),
            _s(getattr(v, "bp_systolic", None)),
            _s(getattr(v, "bp_diastolic", None)),
            _s(getattr(v, "pulse", None)),
            _s(getattr(v, "rr", None)),
            _s(getattr(v, "temp_c", None)),
            _s(getattr(v, "spo2", None)),
            _truncate(getattr(v, "notes", None), 2000) if include_notes else "—",
            _dt_ist(getattr(v, "created_at", None)),
        ])
    ws_vitals = wb.create_sheet()
    _write_sheet(ws_vitals, "Vitals", vit_headers, vit_rows)

    # 4) Prescriptions (header)
    rx_headers = [
        "Visit ID",
        "OP No / Episode",
        "Patient UHID",
        "Patient Name",
        "Signed At (IST)",
        "Signed By",
        "Notes",
    ]
    rx_rows: List[List[Any]] = []
    for v in visits:
        rx = rx_map.get(v.id)
        p = getattr(v, "patient", None)
        rx_rows.append([
            v.id,
            _s(getattr(v, "episode_id", None)),
            _patient_uid(p),
            _patient_name(p),
            _dt_ist(getattr(rx, "signed_at", None)) if rx else "—",
            _doctor_name(getattr(rx, "signer", None)) if rx else "—",
            _truncate(getattr(rx, "notes", None), 2000) if (rx and include_notes) else "—",
        ])
    ws_rx = wb.create_sheet()
    _write_sheet(ws_rx, "Prescriptions", rx_headers, rx_rows)

    # 5) Prescription Items
    rxi_headers = [
        "Visit ID",
        "OP No / Episode",
        "Drug Name",
        "Strength",
        "Frequency",
        "Duration Days",
        "Quantity",
        "Unit Price",
    ]
    rxi_rows: List[List[Any]] = []
    for v in visits:
        items = rx_items_by_visit.get(v.id) or []
        if not items:
            continue
        for it in items:
            rxi_rows.append([
                v.id,
                _s(getattr(v, "episode_id", None)),
                _s(getattr(it, "drug_name", None)),
                _s(getattr(it, "strength", None)),
                _s(getattr(it, "frequency", None)),
                int(getattr(it, "duration_days", 0) or 0),
                int(getattr(it, "quantity", 0) or 0),
                float(getattr(it, "unit_price", 0) or 0),
            ])
    ws_rxi = wb.create_sheet()
    _write_sheet(ws_rxi, "Rx Items", rxi_headers, rxi_rows)

    # 6) Lab Orders
    lab_headers = ["Visit ID", "OP No / Episode", "Test", "Status", "Ordered At (IST)"]
    lab_rows: List[List[Any]] = []
    if visit_ids:
        # we already loaded lab orders above; pull again for detail rows
        lab_orders = (
            db.query(LabOrder)
            .options(selectinload(LabOrder.test))
            .filter(LabOrder.visit_id.in_(visit_ids))
            .order_by(LabOrder.ordered_at.asc())
            .all()
        )
        for o in lab_orders:
            t = getattr(o, "test", None)
            test_label = _s(getattr(t, "name", None) or getattr(t, "code", None))
            vv = next((x for x in visits if x.id == o.visit_id), None)
            lab_rows.append([
                o.visit_id,
                _s(getattr(vv, "episode_id", None)) if vv else "—",
                test_label,
                _s(getattr(o, "status", None)),
                _dt_ist(getattr(o, "ordered_at", None)),
            ])
    ws_lab = wb.create_sheet()
    _write_sheet(ws_lab, "Lab Orders", lab_headers, lab_rows)

    # 7) Radiology Orders
    ris_headers = ["Visit ID", "OP No / Episode", "Test", "Modality", "Status", "Ordered At (IST)"]
    ris_rows: List[List[Any]] = []
    if visit_ids:
        ris_orders = (
            db.query(RadiologyOrder)
            .options(selectinload(RadiologyOrder.test))
            .filter(RadiologyOrder.visit_id.in_(visit_ids))
            .order_by(RadiologyOrder.ordered_at.asc())
            .all()
        )
        for o in ris_orders:
            t = getattr(o, "test", None)
            vv = next((x for x in visits if x.id == o.visit_id), None)
            ris_rows.append([
                o.visit_id,
                _s(getattr(vv, "episode_id", None)) if vv else "—",
                _s(getattr(t, "name", None) or getattr(t, "code", None)),
                _s(getattr(t, "modality", None)),
                _s(getattr(o, "status", None)),
                _dt_ist(getattr(o, "ordered_at", None)),
            ])
    ws_ris = wb.create_sheet()
    _write_sheet(ws_ris, "Radiology Orders", ris_headers, ris_rows)

    # 8) Followups
    fu_headers = [
        "Source Visit ID",
        "OP No / Episode",
        "Patient UHID",
        "Patient Name",
        "Doctor",
        "Department",
        "Due Date",
        "Status",
        "Appointment Linked ID",
        "Note",
        "Created At (IST)",
        "Updated At (IST)",
    ]
    fu_rows: List[List[Any]] = []
    for fu in followups:
        # find visit for episode / patient
        vv = next((x for x in visits if x.id == fu.source_visit_id), None)
        p = getattr(fu, "patient", None) or getattr(vv, "patient", None)
        fu_rows.append([
            fu.source_visit_id,
            _s(getattr(vv, "episode_id", None)) if vv else "—",
            _patient_uid(p),
            _patient_name(p),
            _doctor_name(getattr(fu, "doctor", None)),
            _dept_name(getattr(fu, "department", None)),
            _s(getattr(fu, "due_date", None)),
            _s(getattr(fu, "status", None)),
            _s(getattr(fu, "appointment_id", None)),
            _truncate(getattr(fu, "note", None), 2000) if include_notes else "—",
            _dt_ist(getattr(fu, "created_at", None)),
            _dt_ist(getattr(fu, "updated_at", None)),
        ])
    ws_fu = wb.create_sheet()
    _write_sheet(ws_fu, "Followups", fu_headers, fu_rows)

    # 9) Schedules
    sch_headers = ["Doctor", "Weekday(0=Mon?)", "Start", "End", "Slot Minutes", "Location", "Active"]
    sch_rows: List[List[Any]] = []
    for s in schedules:
        sch_rows.append([
            _doctor_name(getattr(s, "doctor", None)),
            _s(getattr(s, "weekday", None)),
            _time_str(getattr(s, "start_time", None)),
            _time_str(getattr(s, "end_time", None)),
            _s(getattr(s, "slot_minutes", None)),
            _s(getattr(s, "location", None)),
            "YES" if getattr(s, "is_active", False) else "NO",
        ])
    ws_sch = wb.create_sheet()
    _write_sheet(ws_sch, "Schedules", sch_headers, sch_rows)

    # Save
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    fname = f"OPD_Report_{date_from.isoformat()}_to_{date_to.isoformat()}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
