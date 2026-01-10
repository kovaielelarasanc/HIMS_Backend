from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.models.pharmacy_inventory import (
    InventoryItem,
    Supplier,
    InventoryLocation,
    ItemBatch,
    StockTransaction,
)

# ============================================================
# TEMPLATE HEADERS (include all important fields + opening stock)
# ============================================================
TEMPLATE_HEADERS = [
    # identity
    "code",
    "name",
    "qr_number",

    # classification
    "item_type",                 # DRUG | CONSUMABLE | EQUIPMENT
    "is_consumable",             # optional legacy (will be synced)
    "is_active",

    # flags
    "lasa_flag",
    "high_alert_flag",
    "requires_double_check",

    # stock metadata
    "unit",
    "pack_size",

    # UOM conversion
    "base_uom",
    "purchase_uom",
    "conversion_factor",

    "reorder_level",
    "max_level",

    # supplier/procurement
    "manufacturer",
    "default_supplier_id",
    "default_supplier_code",     # helper (resolve supplier)
    "procurement_date",

    # storage
    "storage_condition",

    # defaults
    "default_tax_percent",
    "default_price",
    "default_mrp",

    # regulatory schedule
    "schedule_system",           # IN_DCA | US_CSA
    "schedule_code",             # H/H1/X/B/C1... or II/III/IV...
    "schedule_notes",

    # prescription status
    "prescription_status",       # RX | OTC | SCHEDULED

    # drug fields
    "generic_name",
    "brand_name",
    "dosage_form",
    "strength",
    "active_ingredients",        # comma/pipe/json list
    "route",
    "therapeutic_class",
    "side_effects",
    "drug_interactions",

    # consumable fields
    "material_type",
    "sterility_status",
    "size_dimensions",
    "intended_use",
    "reusable_status",

    # codes
    "atc_code",
    "hsn_code",

    # ✅ opening stock (optional)
    "opening_location_code",     # e.g. MAIN
    "opening_batch_no",          # e.g. BATCH001 (optional)
    "opening_mfg_date",          # YYYY-MM-DD
    "opening_expiry_date",       # YYYY-MM-DD
    "opening_qty",               # number
    "opening_unit_cost",         # number
    "opening_mrp",               # number
    "opening_tax_percent",       # number
    "opening_is_saleable",       # TRUE/FALSE
]

REQUIRED_HEADERS = ["code", "name"]

# accept older/non-DB names and map to template headers
HEADER_ALIASES = {
    # identity
    "item_code": "code",
    "itemcode": "code",
    "medicine_code": "code",
    "drug_code": "code",
    "item_name": "name",
    "medicine_name": "name",
    "barcode": "qr_number",
    "bar_code": "qr_number",
    "qr": "qr_number",

    # misc
    "active": "is_active",

    # schedule
    "schedule": "schedule_code",
    "drug_schedule": "schedule_code",

    # stock
    "min_stock": "reorder_level",
    "minimum_stock": "reorder_level",
    "max_stock": "max_level",
    "maximum_stock": "max_level",

    # tax/pricing
    "gst": "default_tax_percent",
    "tax": "default_tax_percent",
    "tax_percent": "default_tax_percent",
    "mrp": "default_mrp",
    "rate": "default_price",
    "purchase_rate": "default_price",

    # drug field aliases
    "form": "dosage_form",
    "route_of_administration": "route",
    "class_name": "therapeutic_class",

    # supplier
    "supplier": "default_supplier_code",
    "supplier_code": "default_supplier_code",
    "default_supplier": "default_supplier_code",

    # opening stock aliases
    "location": "opening_location_code",
    "location_code": "opening_location_code",
    "opening_location": "opening_location_code",
    "batch": "opening_batch_no",
    "batch_no": "opening_batch_no",
    "qty": "opening_qty",
    "opening_stock": "opening_qty",
    "opening_quantity": "opening_qty",
}

NA_SET = {"", "-", "na", "n/a", "null", "none", "nil"}

SCHEDULE_IN_RE = re.compile(r"^[A-Z0-9]{1,6}$")          # H, H1, X, B, C1...
SCHEDULE_US_RE = re.compile(r"^(I|II|III|IV|V|VI)$")     # II, III, IV...


# ============================================================
# Errors
# ============================================================
@dataclass
class UploadError:
    row: int
    code: Optional[str]
    column: Optional[str]
    message: str


# ============================================================
# Helpers
# ============================================================
def _norm_header(h: Any) -> str:
    s = ("" if h is None else str(h)).strip().lower()
    s = s.replace("\ufeff", "")
    s = re.sub(r"\s+", "_", s)
    s = s.replace("%", "percent")
    return HEADER_ALIASES.get(s, s)


def _safe_text(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    if s.lower() in NA_SET:
        return None
    return s


def _parse_bool(v: Any) -> Optional[bool]:
    s = _safe_text(v)
    if s is None:
        return None
    s2 = s.lower()
    if s2 in {"1", "true", "t", "yes", "y"}:
        return True
    if s2 in {"0", "false", "f", "no", "n"}:
        return False
    return None


def _parse_int(v: Any) -> Optional[int]:
    s = _safe_text(v)
    if s is None:
        return None
    try:
        return int(s)
    except Exception as e:
        raise ValueError(f"Invalid integer '{v}'") from e


def _parse_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()

    s = _safe_text(v)
    if s is None:
        return None

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass

    raise ValueError(f"Invalid date '{v}' (use YYYY-MM-DD or DD-MM-YYYY)")


def _parse_decimal(v: Any) -> Optional[Decimal]:
    if v is None:
        return None
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))

    s = str(v).strip()
    if s.lower() in NA_SET:
        return None

    s = s.replace(",", "").strip()
    if s.endswith("%"):
        s = s[:-1].strip()
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1].strip()

    try:
        return Decimal(s)
    except InvalidOperation as e:
        raise ValueError(f"Invalid number '{v}'") from e


def _parse_list(v: Any) -> Optional[List[str]]:
    s = _safe_text(v)
    if s is None:
        return None

    if s.strip().startswith("["):
        try:
            obj = json.loads(s)
            if isinstance(obj, list):
                out = [str(x).strip() for x in obj if str(x).strip()]
                return out or None
        except Exception:
            pass

    parts = re.split(r"[,\|]", s)
    parts = [p.strip() for p in parts if p.strip()]
    return parts or None


def _norm_item_type(v: Any) -> str:
    s = (_safe_text(v) or "DRUG").upper()
    if s in {"MED", "MEDICINE"}:
        return "DRUG"
    if s in {"CONSUMABLES"}:
        return "CONSUMABLE"
    if s in {"DEVICE"}:
        return "EQUIPMENT"
    return s


def _norm_storage(v: Any) -> str:
    s = (_safe_text(v) or "ROOM_TEMP").upper()
    s = re.sub(r"\s+", "_", s)
    return s


def _norm_schedule_system(v: Any) -> str:
    s = (_safe_text(v) or "IN_DCA").upper()
    if s in {"IN", "INDIA", "DCA"}:
        return "IN_DCA"
    if s in {"US", "USA", "CSA"}:
        return "US_CSA"
    return s


def _norm_schedule_code(v: Any) -> str:
    s = (_safe_text(v) or "").upper()
    s = s.replace(" ", "").replace("-", "").replace("_", "")
    return s


def _norm_ps(v: Any) -> str:
    s = (_safe_text(v) or "RX").upper()
    if s == "SCHEDULE":
        return "SCHEDULED"
    if s in {"PRESCRIPTION"}:
        return "RX"
    if s in {"NONRX", "NON_RX"}:
        return "OTC"
    return s


def _expiry_key(d: Optional[date]) -> int:
    if not d:
        return 0
    return int(d.strftime("%Y%m%d"))


def _validate_required_headers(rows: List[Dict[str, Any]]) -> None:
    if not rows:
        return
    # best-effort: check from first row keys
    keys = set(rows[0].keys())
    missing = [h for h in REQUIRED_HEADERS if h not in keys]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")


# ============================================================
# Parse upload
# ============================================================
def parse_upload_to_rows(filename: str, content_type: str, raw: bytes) -> Tuple[str, List[Dict[str, Any]]]:
    """
    Supports: CSV, XLSX, XLSM
    Returns: (file_type, rows_as_dicts)
    """
    name = (filename or "").lower()

    is_excel = (
        name.endswith(".xlsx")
        or name.endswith(".xlsm")
        or content_type
        in {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel.sheet.macroEnabled.12",
        }
    )

    if is_excel:
        try:
            from openpyxl import load_workbook
        except Exception as e:
            raise ValueError("openpyxl is required for Excel uploads. Install: pip install openpyxl") from e

        keep_vba = name.endswith(".xlsm") or content_type == "application/vnd.ms-excel.sheet.macroEnabled.12"
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True, keep_vba=keep_vba)
        ws = wb.active

        data = list(ws.iter_rows(values_only=True))
        if not data:
            return ("xlsx", [])

        headers = [_norm_header(h) for h in data[0] if h is not None]
        out: List[Dict[str, Any]] = []
        for i, row in enumerate(data[1:], start=2):
            d: Dict[str, Any] = {}
            for j, h in enumerate(headers):
                if not h:
                    continue
                d[h] = row[j] if j < len(row) else None
            # drop fully empty rows
            if any(_safe_text(v) is not None for v in d.values()):
                out.append(d)

        _validate_required_headers(out)
        return ("xlsm" if keep_vba else "xlsx", out)

    # CSV fallback
    try:
        text = raw.decode("utf-8-sig")
    except Exception:
        text = raw.decode("latin-1", errors="replace")

    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
        delim = dialect.delimiter
    except Exception:
        delim = ","

    reader = csv.DictReader(StringIO(text), delimiter=delim)
    out: List[Dict[str, Any]] = []
    for row in reader:
        d = {_norm_header(k): v for k, v in (row or {}).items()}
        if any(_safe_text(v) is not None for v in d.values()):
            out.append(d)

    _validate_required_headers(out)
    return ("csv", out)


# ============================================================
# Validate + Normalize rows
# ============================================================
def validate_item_rows(rows: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[UploadError]]:
    errors: List[UploadError] = []
    normalized: List[Dict[str, Any]] = []
    seen_codes = set()

    for idx, row in enumerate(rows, start=2):
        code = _safe_text(row.get("code"))
        name = _safe_text(row.get("name"))

        if not code:
            continue

        code = code.upper().strip()

        if code in seen_codes:
            errors.append(UploadError(idx, code, "code", "Duplicate code in uploaded file"))
            continue
        seen_codes.add(code)

        if not name:
            errors.append(UploadError(idx, code, "name", "Name is required"))
            continue

        def dec(col: str) -> Optional[Decimal]:
            try:
                return _parse_decimal(row.get(col))
            except ValueError as e:
                errors.append(UploadError(idx, code, col, str(e)))
                return None

        def boo(col: str) -> Optional[bool]:
            rawv = row.get(col)
            b = _parse_bool(rawv)
            if _safe_text(rawv) is not None and b is None:
                errors.append(UploadError(idx, code, col, f"Invalid boolean '{rawv}' (use TRUE/FALSE/1/0)"))
            return b

        def dt(col: str) -> Optional[date]:
            try:
                return _parse_date(row.get(col))
            except ValueError as e:
                errors.append(UploadError(idx, code, col, str(e)))
                return None

        item_type = _norm_item_type(row.get("item_type"))
        is_cons_in = boo("is_consumable")
        is_consumable = (item_type == "CONSUMABLE") if is_cons_in is None else bool(is_cons_in)

        schedule_system = _norm_schedule_system(row.get("schedule_system"))
        schedule_code = _norm_schedule_code(row.get("schedule_code"))
        prescription_status = _norm_ps(row.get("prescription_status"))

        # schedule logic
        if schedule_code in ("RX", "OTC"):
            prescription_status = schedule_code
            schedule_code = ""
        elif schedule_code:
            prescription_status = "SCHEDULED"

            # validate format by system
            if schedule_system == "US_CSA":
                if not SCHEDULE_US_RE.match(schedule_code):
                    errors.append(UploadError(idx, code, "schedule_code", "Invalid US_CSA schedule_code (II/III/IV/V etc.)"))
            else:
                if not SCHEDULE_IN_RE.match(schedule_code):
                    errors.append(UploadError(idx, code, "schedule_code", "Invalid IN_DCA schedule_code (H, H1, X, B, C1...)"))

        if prescription_status == "SCHEDULED" and not schedule_code:
            errors.append(UploadError(idx, code, "schedule_code", "schedule_code is required when prescription_status is SCHEDULED"))

        opening_qty = dec("opening_qty")
        if opening_qty is not None and opening_qty < 0:
            errors.append(UploadError(idx, code, "opening_qty", "opening_qty must be >= 0"))

        nrow: Dict[str, Any] = {
            "code": code,
            "name": name.strip(),
            "qr_number": _safe_text(row.get("qr_number")),

            "item_type": item_type,
            "is_consumable": bool(is_consumable),
            "is_active": True if boo("is_active") is None else bool(boo("is_active")),

            "lasa_flag": bool(boo("lasa_flag") or False),
            "high_alert_flag": bool(boo("high_alert_flag") or False),
            "requires_double_check": bool(boo("requires_double_check") or False),

            "unit": _safe_text(row.get("unit")) or "unit",
            "pack_size": _safe_text(row.get("pack_size")) or "1",

            "base_uom": _safe_text(row.get("base_uom")) or "unit",
            "purchase_uom": _safe_text(row.get("purchase_uom")) or "unit",
            "conversion_factor": dec("conversion_factor") or Decimal("1"),

            "reorder_level": dec("reorder_level") or Decimal("0"),
            "max_level": dec("max_level") or Decimal("0"),

            "manufacturer": _safe_text(row.get("manufacturer")) or "",
            "default_supplier_id": None,
            "default_supplier_code": _safe_text(row.get("default_supplier_code")),
            "procurement_date": dt("procurement_date"),

            "storage_condition": _norm_storage(row.get("storage_condition")),

            "default_tax_percent": dec("default_tax_percent") or Decimal("0"),
            "default_price": dec("default_price") or Decimal("0"),
            "default_mrp": dec("default_mrp") or Decimal("0"),

            "schedule_system": schedule_system,
            "schedule_code": schedule_code or "",
            "schedule_notes": _safe_text(row.get("schedule_notes")) or "",

            "prescription_status": prescription_status,

            "generic_name": _safe_text(row.get("generic_name")) or "",
            "brand_name": _safe_text(row.get("brand_name")) or "",
            "dosage_form": _safe_text(row.get("dosage_form")) or "",
            "strength": _safe_text(row.get("strength")) or "",
            "active_ingredients": _parse_list(row.get("active_ingredients")),
            "route": _safe_text(row.get("route")) or "",
            "therapeutic_class": _safe_text(row.get("therapeutic_class")) or "",
            "side_effects": _safe_text(row.get("side_effects")) or "",
            "drug_interactions": _safe_text(row.get("drug_interactions")) or "",

            "material_type": _safe_text(row.get("material_type")) or "",
            "sterility_status": _safe_text(row.get("sterility_status")) or "",
            "size_dimensions": _safe_text(row.get("size_dimensions")) or "",
            "intended_use": _safe_text(row.get("intended_use")) or "",
            "reusable_status": _safe_text(row.get("reusable_status")) or "",

            "atc_code": _safe_text(row.get("atc_code")) or "",
            "hsn_code": _safe_text(row.get("hsn_code")) or "",

            # opening stock
            "opening_location_code": (_safe_text(row.get("opening_location_code")) or "MAIN").upper(),
            "opening_batch_no": _safe_text(row.get("opening_batch_no")),
            "opening_mfg_date": dt("opening_mfg_date"),
            "opening_expiry_date": dt("opening_expiry_date"),
            "opening_qty": opening_qty,
            "opening_unit_cost": dec("opening_unit_cost"),
            "opening_mrp": dec("opening_mrp"),
            "opening_tax_percent": dec("opening_tax_percent"),
            "opening_is_saleable": True if boo("opening_is_saleable") is None else bool(boo("opening_is_saleable")),
        }

        # supplier id parse
        try:
            nrow["default_supplier_id"] = _parse_int(row.get("default_supplier_id"))
        except ValueError as e:
            errors.append(UploadError(idx, code, "default_supplier_id", str(e)))
            nrow["default_supplier_id"] = None

        # enforce high alert rule
        if nrow["high_alert_flag"] and not nrow["requires_double_check"]:
            nrow["requires_double_check"] = True

        # max_level >= reorder_level
        if nrow["max_level"] < nrow["reorder_level"]:
            errors.append(UploadError(idx, code, "max_level", "max_level must be >= reorder_level"))

        normalized.append(nrow)

    return normalized, errors


# ============================================================
# Stock creation / adjustment
# ============================================================
def _get_location(db: Session, code: str) -> Optional[InventoryLocation]:
    return db.query(InventoryLocation).filter(InventoryLocation.code == code).first()


def _get_or_create_location(db: Session, code: str) -> InventoryLocation:
    loc = _get_location(db, code)
    if loc:
        return loc
    loc = InventoryLocation(code=code, name=code, is_active=True)
    db.add(loc)
    db.flush()
    return loc


def _get_or_create_batch(
    db: Session,
    *,
    item_id: int,
    location_id: int,
    batch_no: str,
    mfg_date: Optional[date],
    expiry_date: Optional[date],
    unit_cost: Decimal,
    mrp: Decimal,
    tax_percent: Decimal,
    is_saleable: bool,
) -> ItemBatch:
    ek = _expiry_key(expiry_date)
    batch = (
        db.query(ItemBatch)
        .filter(
            ItemBatch.item_id == item_id,
            ItemBatch.location_id == location_id,
            ItemBatch.batch_no == batch_no,
            ItemBatch.expiry_key == ek,
        )
        .first()
    )
    if batch:
        # keep latest meta if user provides
        batch.mfg_date = mfg_date or batch.mfg_date
        batch.expiry_date = expiry_date or batch.expiry_date
        batch.expiry_key = _expiry_key(batch.expiry_date)
        batch.unit_cost = unit_cost if unit_cost is not None else batch.unit_cost
        batch.mrp = mrp if mrp is not None else batch.mrp
        batch.tax_percent = tax_percent if tax_percent is not None else batch.tax_percent
        batch.is_saleable = bool(is_saleable)
        return batch

    batch = ItemBatch(
        item_id=item_id,
        location_id=location_id,
        batch_no=batch_no,
        mfg_date=mfg_date,
        expiry_date=expiry_date,
        expiry_key=_expiry_key(expiry_date),
        current_qty=Decimal("0"),
        reserved_qty=Decimal("0"),
        unit_cost=unit_cost,
        mrp=mrp,
        tax_percent=tax_percent,
        is_active=True,
        is_saleable=bool(is_saleable),
    )
    db.add(batch)
    db.flush()
    return batch


def _apply_opening_stock(
    db: Session,
    *,
    item: InventoryItem,
    row: Dict[str, Any],
    user_id: Optional[int],
    create_missing_locations: bool,
) -> None:
    qty: Optional[Decimal] = row.get("opening_qty")
    if qty is None:
        return

    loc_code = (row.get("opening_location_code") or "MAIN").upper()
    loc = _get_or_create_location(db, loc_code) if create_missing_locations else _get_location(db, loc_code)
    if not loc:
        raise ValueError(f"Opening stock location '{loc_code}' not found")

    batch_no = row.get("opening_batch_no") or f"OPEN-{item.code}"
    mfg_date = row.get("opening_mfg_date")
    expiry_date = row.get("opening_expiry_date")

    unit_cost = row.get("opening_unit_cost") or Decimal("0")
    mrp = row.get("opening_mrp") or (row.get("default_mrp") or Decimal("0"))
    tax_percent = row.get("opening_tax_percent") or (row.get("default_tax_percent") or Decimal("0"))
    is_saleable = bool(row.get("opening_is_saleable", True))

    batch = _get_or_create_batch(
        db,
        item_id=item.id,
        location_id=loc.id,
        batch_no=batch_no,
        mfg_date=mfg_date,
        expiry_date=expiry_date,
        unit_cost=unit_cost,
        mrp=mrp,
        tax_percent=tax_percent,
        is_saleable=is_saleable,
    )

    # We SET batch current_qty to the uploaded opening_qty (not add)
    current = Decimal(str(batch.current_qty or 0))
    target = Decimal(str(qty))
    delta = target - current

    if delta == 0:
        return

    batch.current_qty = target

    txn = StockTransaction(
        location_id=loc.id,
        item_id=item.id,
        batch_id=batch.id,
        txn_time=datetime.utcnow(),
        txn_type="OPENING",
        ref_type="BULK_UPLOAD",
        ref_id=None,
        ref_line_id=None,
        quantity_change=delta,
        unit_cost=unit_cost,
        mrp=mrp,
        remark=f"Opening stock set to {target} via bulk upload",
        user_id=user_id,
        patient_id=None,
        visit_id=None,
        doctor_id=None,
    )
    db.add(txn)


# ============================================================
# Commit to DB
# ============================================================
def apply_items_import(
    db: Session,
    normalized_rows: List[Dict[str, Any]],
    *,
    update_blanks: bool = False,
    create_missing_locations: bool = True,
    user_id: Optional[int] = None,
) -> Tuple[int, int, int, List[UploadError]]:
    created = 0
    updated = 0
    skipped = 0
    errors: List[UploadError] = []

    if not normalized_rows:
        return 0, 0, 0, []

    codes = [r["code"] for r in normalized_rows]
    existing_items = db.query(InventoryItem).filter(InventoryItem.code.in_(codes)).all()
    existing_by_code = {it.code: it for it in existing_items}

    # QR uniqueness check
    qrs = [r.get("qr_number") for r in normalized_rows if r.get("qr_number")]
    if qrs:
        qr_existing = db.query(InventoryItem).filter(InventoryItem.qr_number.in_(qrs)).all()
        qr_to_code = {it.qr_number: it.code for it in qr_existing if it.qr_number}
        for row_idx, r in enumerate(normalized_rows, start=2):
            qr = r.get("qr_number")
            if qr and qr in qr_to_code and qr_to_code[qr] != r["code"]:
                errors.append(UploadError(row_idx, r["code"], "qr_number", f"QR already used by item '{qr_to_code[qr]}'"))

    if errors:
        return 0, 0, 0, errors

    # supplier resolution
    supplier_keys = { (r.get("default_supplier_code") or "").strip() for r in normalized_rows if r.get("default_supplier_code") }
    supplier_keys = {k for k in supplier_keys if k}
    supplier_by_code: Dict[str, Supplier] = {}
    supplier_by_name: Dict[str, Supplier] = {}

    if supplier_keys:
        sups = (
            db.query(Supplier)
            .filter(or_(Supplier.code.in_(supplier_keys), Supplier.name.in_(supplier_keys)))
            .all()
        )
        for s in sups:
            if getattr(s, "code", None):
                supplier_by_code[str(s.code).strip().upper()] = s
            if getattr(s, "name", None):
                supplier_by_name[str(s.name).strip().lower()] = s

        for row_idx, r in enumerate(normalized_rows, start=2):
            if r.get("default_supplier_id"):
                continue
            key = (r.get("default_supplier_code") or "").strip()
            if not key:
                continue
            sup = supplier_by_code.get(key.upper()) or supplier_by_name.get(key.lower())
            if not sup:
                errors.append(UploadError(row_idx, r["code"], "default_supplier_code", f"Supplier '{key}' not found"))
            else:
                r["default_supplier_id"] = int(sup.id)

    if errors:
        return 0, 0, 0, errors

    def should_set(v: Any) -> bool:
        if v is None:
            return False
        if isinstance(v, str) and v.strip() == "" and not update_blanks:
            return False
        if isinstance(v, list) and len(v) == 0 and not update_blanks:
            return False
        return True

    # helper-only keys (not DB columns)
    helper_keys = {
        "default_supplier_code",
        "opening_location_code",
        "opening_batch_no",
        "opening_mfg_date",
        "opening_expiry_date",
        "opening_qty",
        "opening_unit_cost",
        "opening_mrp",
        "opening_tax_percent",
        "opening_is_saleable",
    }

    try:
        for row_idx, r0 in enumerate(normalized_rows, start=2):
            r = dict(r0)
            code = r["code"]
            existing = existing_by_code.get(code)

            # strip helper keys for InventoryItem create/update
            db_payload = {k: v for k, v in r.items() if k not in helper_keys}

            if existing:
                # UPDATE
                for field, val in db_payload.items():
                    if field == "code":
                        continue
                    if should_set(val):
                        setattr(existing, field, val)

                # enforce schedule consistency
                sc = (getattr(existing, "schedule_code", "") or "").strip()
                ps = (getattr(existing, "prescription_status", "RX") or "RX").upper()
                if sc and ps != "SCHEDULED":
                    existing.prescription_status = "SCHEDULED"
                if ps == "SCHEDULED" and not sc:
                    raise ValueError(f"Row {row_idx} ({code}): schedule_code required for SCHEDULED")

                updated += 1

                # opening stock apply
                _apply_opening_stock(
                    db,
                    item=existing,
                    row=r,
                    user_id=user_id,
                    create_missing_locations=create_missing_locations,
                )

            else:
                # CREATE
                item = InventoryItem(**db_payload)
                db.add(item)
                db.flush()

                if not item.qr_number:
                    item.qr_number = f"MED-{item.id:06d}"

                created += 1

                _apply_opening_stock(
                    db,
                    item=item,
                    row=r,
                    user_id=user_id,
                    create_missing_locations=create_missing_locations,
                )

        db.commit()
        return created, updated, skipped, []

    except IntegrityError as e:
        db.rollback()
        return 0, 0, 0, [UploadError(row=0, code=None, column=None, message=f"DB constraint error: {str(e.orig)}")]
    except Exception as e:
        db.rollback()
        return 0, 0, 0, [UploadError(row=0, code=None, column=None, message=f"Unexpected error: {str(e)}")]


# ============================================================
# Template Generators
# ============================================================
def make_csv_template_bytes() -> bytes:
    output = StringIO()
    w = csv.writer(output)
    w.writerow(TEMPLATE_HEADERS)

    # sample DRUG row (includes opening stock)
    w.writerow([
        "ITEM001", "Paracetamol 500mg", "MED-000001",
        "DRUG", "FALSE", "TRUE",
        "FALSE", "FALSE", "FALSE",
        "TAB", "10",
        "TAB", "BOX", "10",
        "50", "500",
        "ABC Pharma", "", "SUP001", "",
        "ROOM_TEMP",
        "12", "1.50", "2.00",
        "IN_DCA", "H", "Rx only",
        "RX",
        "Paracetamol", "", "Tablet", "500mg", "Paracetamol", "oral", "Analgesic", "", "",
        "", "", "", "", "",
        "N02BE01", "3004",
        "MAIN", "BATCH001", "", "2028-12-31", "100", "1.00", "2.00", "12", "TRUE",
    ])

    data = output.getvalue().encode("utf-8-sig")
    return data


def make_excel_template_bytes(*, macro_enabled: bool = False, base_xlsm_path: Optional[Path] = None) -> Tuple[bytes, str]:
    """
    Returns (bytes, file_ext)
    - If macro_enabled=True and base_xlsm_path exists => load it with keep_vba and write headers/samples => real XLSM
    - Else => generate XLSX
    """
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
    except Exception as e:
        raise ValueError("openpyxl is required for Excel template. Install: pip install openpyxl") from e

    if macro_enabled and base_xlsm_path and base_xlsm_path.exists():
        wb = load_workbook(str(base_xlsm_path), keep_vba=True)
        ws = wb.active
        ws.title = "Items"
        file_ext = "xlsm"
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Items"
        file_ext = "xlsx"

    ws.delete_rows(1, ws.max_row)
    ws.append(TEMPLATE_HEADERS)

    # styling
    header_font = Font(bold=True)
    for col in range(1, len(TEMPLATE_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = 18

    ws.freeze_panes = "A2"

    # dropdown validations
    def add_list_dv(col_letter: str, values: List[str]):
        dv = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}2000")

    headers_to_col = {h: i + 1 for i, h in enumerate(TEMPLATE_HEADERS)}
    def col_letter(h: str) -> str:
        return ws.cell(row=1, column=headers_to_col[h]).column_letter

    add_list_dv(col_letter("item_type"), ["DRUG", "CONSUMABLE", "EQUIPMENT"])
    add_list_dv(col_letter("is_active"), ["TRUE", "FALSE"])
    add_list_dv(col_letter("lasa_flag"), ["TRUE", "FALSE"])
    add_list_dv(col_letter("high_alert_flag"), ["TRUE", "FALSE"])
    add_list_dv(col_letter("requires_double_check"), ["TRUE", "FALSE"])
    add_list_dv(col_letter("schedule_system"), ["IN_DCA", "US_CSA"])
    add_list_dv(col_letter("prescription_status"), ["RX", "OTC", "SCHEDULED"])
    add_list_dv(col_letter("opening_is_saleable"), ["TRUE", "FALSE"])
    add_list_dv(col_letter("storage_condition"), ["ROOM_TEMP", "COLD_CHAIN", "FROZEN", "CONTROLLED"])

    # sample rows
    sample = {h: "" for h in TEMPLATE_HEADERS}
    sample.update({
        "code": "ITEM001",
        "name": "Paracetamol 500mg",
        "qr_number": "",
        "item_type": "DRUG",
        "is_active": "TRUE",
        "lasa_flag": "FALSE",
        "high_alert_flag": "FALSE",
        "requires_double_check": "FALSE",
        "unit": "TAB",
        "pack_size": "10",
        "base_uom": "TAB",
        "purchase_uom": "BOX",
        "conversion_factor": "10",
        "reorder_level": "50",
        "max_level": "500",
        "manufacturer": "ABC Pharma",
        "default_supplier_code": "SUP001",
        "storage_condition": "ROOM_TEMP",
        "default_tax_percent": "12",
        "default_price": "1.50",
        "default_mrp": "2.00",
        "schedule_system": "IN_DCA",
        "schedule_code": "H",
        "schedule_notes": "Rx only",
        "prescription_status": "RX",
        "generic_name": "Paracetamol",
        "dosage_form": "Tablet",
        "strength": "500mg",
        "active_ingredients": "Paracetamol",
        "route": "oral",
        "therapeutic_class": "Analgesic",
        "hsn_code": "3004",
        "opening_location_code": "MAIN",
        "opening_batch_no": "BATCH001",
        "opening_expiry_date": "2028-12-31",
        "opening_qty": "100",
        "opening_unit_cost": "1.00",
        "opening_mrp": "2.00",
        "opening_tax_percent": "12",
        "opening_is_saleable": "TRUE",
    })
    ws.append([sample.get(h, "") for h in TEMPLATE_HEADERS])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), file_ext
