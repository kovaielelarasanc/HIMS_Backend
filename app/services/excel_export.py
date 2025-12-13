from __future__ import annotations

from datetime import date
from decimal import Decimal
from typing import Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _money(x) -> float:
    try:
        return float(Decimal(str(x or "0")))
    except Exception:
        return 0.0


def build_supplier_ledger_excel(fp, invoices: Iterable):
    wb = Workbook()
    ws = wb.active
    ws.title = "Supplier Ledger"

    headers = [
        "GRN No", "Invoice No", "Invoice Date", "Supplier",
        "Invoice Amount", "Paid Amount", "Outstanding",
        "Status", "Overdue", "Last Payment Date",
    ]
    ws.append(headers)

    for inv in invoices:
        ws.append([
            getattr(inv, "grn_number", ""),
            getattr(inv, "invoice_number", ""),
            getattr(inv, "invoice_date", None),
            getattr(getattr(inv, "supplier", None), "name", "") or "",
            _money(getattr(inv, "invoice_amount", 0)),
            _money(getattr(inv, "paid_amount", 0)),
            _money(getattr(inv, "outstanding_amount", 0)),
            getattr(inv, "status", ""),
            "YES" if getattr(inv, "is_overdue", False) else "NO",
            getattr(inv, "last_payment_date", None),
        ])

    # autosize
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(fp)


def build_supplier_monthly_summary_excel(fp, summary):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Summary {summary.month}"

    headers = ["Supplier ID", "Month", "Total Purchase", "Total Paid", "Pending", "Overdue Invoices", "Last Payment Date"]
    ws.append(headers)

    for r in summary.rows:
        ws.append([
            r.supplier_id,
            r.month,
            _money(r.total_purchase),
            _money(r.total_paid),
            _money(r.pending_amount),
            int(r.overdue_invoices),
            r.last_payment_date,
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(fp)
