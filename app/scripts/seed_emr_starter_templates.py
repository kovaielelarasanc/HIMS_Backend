# FILE: app/seed/seed_emr_starter_templates.py
from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from fastapi import HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import func

from openpyxl import load_workbook

from app.models.emr_all import (
    EmrDepartment, EmrRecordType,
    EmrTemplate, EmrTemplateVersion, EmrTemplateStatus,
    EmrSectionLibrary
)

# Block library is optional (template schema can still reference block codes)
try:
    from app.models.emr_template_library import EmrTemplateBlock
except Exception:
    EmrTemplateBlock = None  # type: ignore


# ----------------------------
# Helpers
# ----------------------------
def _norm(v: str) -> str:
    return (v or "").strip().upper().replace(" ", "_")


def _dumps(x: Any) -> str:
    return json.dumps(x, ensure_ascii=False)


def _loads(s: Optional[str], default):
    try:
        return json.loads(s or "")
    except Exception:
        return default


def _safe_name(name: str, max_len: int = 120) -> str:
    n = (name or "").strip()
    if len(n) <= max_len:
        return n
    return n[:max_len].rstrip()


def _safe_desc(desc: Optional[str], max_len: int = 800) -> Optional[str]:
    if not desc:
        return None
    d = str(desc).strip()
    if not d:
        return None
    return d[:max_len].rstrip()


def _ensure_record_types(db: Session, create_missing: bool = False) -> None:
    """
    Ensures core record types exist.
    """
    core = [
        ("OPD_NOTE", "OPD Note", "Clinical", 10),
        ("IPD_NOTE", "IPD Note", "Clinical", 20),
        ("DISCHARGE_SUMMARY", "Discharge Summary", "Clinical", 30),
        ("NURSING_NOTE", "Nursing Note", "Nursing", 40),
        ("ANC_NOTE", "ANC Note", "OBGYN", 50),
        ("OT_NOTE", "OT Note", "Surgery", 60),
    ]
    existing = {r.code for r in db.query(EmrRecordType).all()}
    for code, label, cat, order in core:
        if code in existing:
            continue
        if not create_missing:
            continue
        db.add(EmrRecordType(
            code=code, label=label, category=cat, is_active=True, display_order=order
        ))
    db.commit()


def _upsert_section(db: Session, *, code: str, label: str, group: Optional[str], display_order: int) -> None:
    code = _norm(code)
    row = db.query(EmrSectionLibrary).filter(func.upper(EmrSectionLibrary.code) == code).one_or_none()
    if row:
        # keep label fresh
        row.label = label.strip()
        row.group = group
        row.display_order = int(display_order)
        row.is_active = True
    else:
        db.add(EmrSectionLibrary(
            code=code,
            label=label.strip(),
            group=group,
            is_active=True,
            display_order=int(display_order),
        ))


def _ensure_core_sections(db: Session) -> None:
    """
    Minimal section library to power wizard + starter templates.
    You can add more freely later.
    """
    sections = [
        ("PATIENT_HEADER", "Patient Header", "COMMON", 10),
        ("VITALS", "Vitals", "COMMON", 20),
        ("SOAP", "SOAP Note", "SOAP", 30),
        ("DIAGNOSIS", "Diagnosis", "SOAP", 40),
        ("MEDICATIONS", "Medications", "SOAP", 50),
        ("INVESTIGATIONS", "Investigations", "SOAP", 60),
        ("ADVICE", "Advice", "SOAP", 70),
        ("SIGNATURES", "Signatures", "COMMON", 90),
        ("ATTACHMENTS", "Attachments", "COMMON", 95),

        ("NURSING_ASSESSMENT", "Nursing Assessment", "NURSING", 110),
        ("PAIN", "Pain Score", "NURSING", 120),
        ("FALL_RISK", "Fall Risk", "NURSING", 130),
        ("BRADEN", "Braden Scale", "NURSING", 140),
        ("INTAKE_OUTPUT", "Intake / Output", "NURSING", 150),

        ("HOSPITAL_COURSE", "Hospital Course", "DISCHARGE", 210),
        ("PROCEDURES", "Procedures", "DISCHARGE", 220),
        ("DISCHARGE_MEDICATIONS", "Discharge Medications", "DISCHARGE", 230),
        ("DISCHARGE_INSTRUCTIONS", "Discharge Instructions", "DISCHARGE", 240),
        ("FOLLOW_UP", "Follow Up", "DISCHARGE", 250),

        ("ANC_PROFILE", "ANC Profile", "OBGYN", 310),
        ("MENSTRUAL_HISTORY", "Menstrual History", "OBGYN", 320),
        ("OB_HISTORY", "OB History (GTPAL)", "OBGYN", 330),
        ("FETAL_ASSESSMENT", "Fetal Assessment", "OBGYN", 340),
        ("PLAN", "Plan", "OBGYN", 350),

        ("OT_CHECKLIST", "OT WHO Checklist", "OT", 410),
        ("ANESTHESIA", "Anaesthesia Record", "OT", 420),
        ("PROCEDURE_NOTE", "Procedure Note", "OT", 430),
        ("COUNTS", "Counts", "OT", 440),
        ("POST_OP_ORDERS", "Post-Op Orders", "OT", 450),
    ]

    for code, label, group, order in sections:
        _upsert_section(db, code=code, label=label, group=group, display_order=order)

    db.commit()


def _choose_record_type(area: str, name: str, who: str) -> str:
    a = (area or "").strip().upper()
    n = (name or "").strip().upper()
    w = (who or "").strip().upper()

    if "DISCHARGE" in n:
        return "DISCHARGE_SUMMARY"
    if "NURS" in n or "NURSE" in w:
        return "NURSING_NOTE"
    if "ANC" in n or "ANTENATAL" in n or "ANC" in a:
        return "ANC_NOTE"
    if "OT" in a or "OT" in n or "PRE-OT" in a or "CATH LAB" in a:
        return "OT_NOTE"
    if "IPD" in a or "ICU" in a or "ER/IPD" in a:
        return "IPD_NOTE"
    if "OPD" in a:
        return "OPD_NOTE"
    return "OPD_NOTE"


def _sections_for(rt: str) -> List[str]:
    rt = _norm(rt)
    if rt == "NURSING_NOTE":
        return ["PATIENT_HEADER", "VITALS", "NURSING_ASSESSMENT", "PAIN", "FALL_RISK", "BRADEN", "INTAKE_OUTPUT", "SIGNATURES"]
    if rt == "DISCHARGE_SUMMARY":
        return ["PATIENT_HEADER", "DIAGNOSIS", "HOSPITAL_COURSE", "PROCEDURES", "DISCHARGE_MEDICATIONS", "DISCHARGE_INSTRUCTIONS", "FOLLOW_UP", "SIGNATURES", "ATTACHMENTS"]
    if rt == "ANC_NOTE":
        return ["PATIENT_HEADER", "ANC_PROFILE", "MENSTRUAL_HISTORY", "OB_HISTORY", "VITALS", "FETAL_ASSESSMENT", "PLAN", "SIGNATURES"]
    if rt == "OT_NOTE":
        return ["PATIENT_HEADER", "OT_CHECKLIST", "ANESTHESIA", "PROCEDURE_NOTE", "COUNTS", "POST_OP_ORDERS", "SIGNATURES", "ATTACHMENTS"]
    if rt == "IPD_NOTE":
        return ["PATIENT_HEADER", "VITALS", "SOAP", "DIAGNOSIS", "MEDICATIONS", "INVESTIGATIONS", "ADVICE", "SIGNATURES", "ATTACHMENTS"]
    return ["PATIENT_HEADER", "VITALS", "SOAP", "DIAGNOSIS", "MEDICATIONS", "INVESTIGATIONS", "ADVICE", "SIGNATURES", "ATTACHMENTS"]


def _schema_for(rt: str, *, area: str, who: str, mandatory: str) -> Dict[str, Any]:
    """
    Frontend-friendly schema:
    - sections[] in order
    - each section has items: block refs + optional free text fields
    """
    rt = _norm(rt)
    sections = _sections_for(rt)

    # block refs (these should exist in your block library seed)
    block_map = {
        "VITALS": ["BLOCK_VITALS"],
        "SOAP": ["BLOCK_SOAP_MINI"],
        "DIAGNOSIS": ["BLOCK_DIAGNOSIS_LIST"],
        "MEDICATIONS": ["BLOCK_MEDICATION_LIST"],
        "SIGNATURES": ["BLOCK_SIGNATURES"],
        "ATTACHMENTS": ["BLOCK_ATTACHMENTS"],

        "NURSING_ASSESSMENT": ["BLOCK_NURSING_ASSESSMENT_BASIC"],
        "PAIN": ["BLOCK_PAIN_SCALE"],
        "FALL_RISK": ["BLOCK_FALL_RISK_SIMPLE"],
        "BRADEN": ["BLOCK_BRADEN_SCALE"],
        "INTAKE_OUTPUT": ["BLOCK_INTAKE_OUTPUT"],

        "PROCEDURES": ["BLOCK_PROCEDURES_TABLE"],
        "DISCHARGE_MEDICATIONS": ["BLOCK_DISCHARGE_MEDICATIONS"],
        "DISCHARGE_INSTRUCTIONS": ["BLOCK_DISCHARGE_INSTRUCTIONS"],

        "ANC_PROFILE": ["BLOCK_ANC_PROFILE"],
        "MENSTRUAL_HISTORY": ["BLOCK_MENSTRUAL_HISTORY"],
        "OB_HISTORY": ["BLOCK_OB_HISTORY_GTPAL"],
        "FETAL_ASSESSMENT": ["BLOCK_FETAL_ASSESSMENT"],

        "OT_CHECKLIST": ["BLOCK_OT_WHO_CHECKLIST"],
        "ANESTHESIA": ["BLOCK_ANESTHESIA_RECORD_BASIC"],
        "PROCEDURE_NOTE": ["BLOCK_PROCEDURES_TABLE"],
    }

    ui_sections = []
    for sc in sections:
        items = [{"type": "block", "code": b} for b in block_map.get(sc, [])]

        # lightweight free-text “doctor notes” where needed
        if sc in ("ADVICE", "HOSPITAL_COURSE", "FOLLOW_UP", "PLAN", "POST_OP_ORDERS"):
            items.append({"type": "field", "field": f"{sc.lower()}_text", "label": "Notes", "ui": "textarea"})

        ui_sections.append({
            "code": sc,
            "title": sc.replace("_", " ").title(),
            "items": items,
        })

    return {
        "v": 1,
        "record_type": rt,
        "clinical": {
            "area": (area or "").strip(),
            "who_fills": (who or "").strip(),
            "mandatory_hint": (mandatory or "").strip(),
        },
        "sections": ui_sections,
        "rules": {
            "require_signature_on_sign": True,
            "warn_if_empty_sections": True,
        }
    }


# ----------------------------
# Excel Parsing (your file format)
# ----------------------------
@dataclass
class XRow:
    dept: str           # "__ALL__" for common table, else dept name
    template_name: str
    area: str
    who: str
    mandatory: str


def _parse_excel(path: str) -> List[XRow]:
    wb = load_workbook(path)
    ws = wb.active

    rows: List[XRow] = []
    mode = "none"
    current_dept: Optional[str] = None

    for r in ws.iter_rows(min_row=1, max_col=5, values_only=True):
        c0, c1, c2, c3, c4 = [(x if x is not None else "") for x in r]
        c0s = str(c0).strip()
        c1s = str(c1).strip()

        # detect section switch
        if c0s.upper().startswith("DEPARTMENT-WISE CASE SHEETS"):
            mode = "deptwise"
            current_dept = None
            continue

        # common header
        if c0s.upper() == "DEPARTMENT" and str(c1).strip().upper() == "TEMPLATE NAME":
            # if we haven’t yet hit deptwise, this is common table header too
            if mode != "deptwise":
                mode = "common"
            continue

        if mode == "common":
            # common table uses "Common (All)" in dept column
            if c1s:
                dept = "__ALL__"
                rows.append(XRow(
                    dept=dept,
                    template_name=c1s,
                    area=str(c2).strip(),
                    who=str(c3).strip(),
                    mandatory=str(c4).strip(),
                ))
            continue

        if mode == "deptwise":
            # dept title line (only first col filled)
            if c0s and not c1s and c0s.upper() != "DEPARTMENT":
                current_dept = c0s
                continue

            # skip internal header inside each dept
            if c0s.upper() == "DEPARTMENT" and c1s.upper() == "TEMPLATE NAME":
                continue

            if current_dept and c1s:
                rows.append(XRow(
                    dept=current_dept,
                    template_name=c1s,
                    area=str(c2).strip(),
                    who=str(c3).strip(),
                    mandatory=str(c4).strip(),
                ))
            continue

    return rows


# ----------------------------
# Template creation
# ----------------------------
def _match_dept_code(db: Session, dept_name_or_code: str) -> Optional[str]:
    s = (dept_name_or_code or "").strip()
    if not s:
        return None
    norm = _norm(s)

    deps = db.query(EmrDepartment).all()

    # 1) exact code
    for d in deps:
        if (d.code or "").upper() == norm:
            return d.code

    # 2) name exact
    for d in deps:
        if (d.name or "").strip().lower() == s.lower():
            return d.code

    # 3) name contains
    for d in deps:
        if s.lower() in (d.name or "").lower():
            return d.code

    # 4) code contains
    for d in deps:
        if norm in (d.code or "").upper():
            return d.code

    return None


def _ensure_dept_if_missing(db: Session, dept_name: str, create_missing: bool) -> Optional[str]:
    code = _match_dept_code(db, dept_name)
    if code:
        return code

    if not create_missing:
        return None

    # create new dept using normalized name
    new_code = _norm(dept_name)[:32]
    exists = db.query(EmrDepartment.id).filter(func.upper(EmrDepartment.code) == new_code).first()
    if exists:
        return new_code

    db.add(EmrDepartment(code=new_code, name=dept_name.strip(), is_active=True, display_order=1000))
    db.commit()
    return new_code


def _ensure_record_type(db: Session, rt: str, create_missing: bool) -> bool:
    rt = _norm(rt)
    row = db.query(EmrRecordType).filter(EmrRecordType.code == rt).one_or_none()
    if row:
        return bool(row.is_active)
    if not create_missing:
        return False
    db.add(EmrRecordType(code=rt, label=rt.replace("_", " ").title(), category="Clinical", is_active=True, display_order=1000))
    db.commit()
    return True


def _template_exists(db: Session, *, dept_code: str, record_type_code: str, name: str) -> Optional[int]:
    tid = db.query(EmrTemplate.id).filter(
        EmrTemplate.dept_code == dept_code,
        EmrTemplate.record_type_code == record_type_code,
        func.lower(EmrTemplate.name) == name.lower(),
    ).scalar()
    return int(tid) if tid else None


def _create_template_with_version(
    db: Session,
    *,
    dept_code: str,
    record_type_code: str,
    name: str,
    description: Optional[str],
    sections: List[str],
    schema_obj: Dict[str, Any],
    publish: bool,
    is_default: bool,
    user_id: int,
) -> int:
    status = EmrTemplateStatus.PUBLISHED if publish else EmrTemplateStatus.DRAFT

    tpl = EmrTemplate(
        dept_code=dept_code,
        record_type_code=record_type_code,
        name=_safe_name(name),
        description=_safe_desc(description),
        restricted=False,
        premium=False,
        is_default=bool(is_default),
        status=status,
        created_by_user_id=user_id,
        updated_by_user_id=user_id,
    )
    db.add(tpl)
    db.flush()

    ver = EmrTemplateVersion(
        template_id=int(tpl.id),
        version_no=1,
        changelog="Seeded starter template",
        sections_json=_dumps(sections),
        schema_json=_dumps(schema_obj),
        created_by_user_id=user_id,
    )
    db.add(ver)
    db.flush()

    tpl.active_version_id = int(ver.id)
    if publish:
        tpl.published_version_id = int(ver.id)

    # keep one default per dept+type
    if tpl.is_default:
        db.query(EmrTemplate).filter(
            EmrTemplate.dept_code == dept_code,
            EmrTemplate.record_type_code == record_type_code,
            EmrTemplate.id != tpl.id,
        ).update({"is_default": False})

    db.commit()
    return int(tpl.id)


def _add_new_version(
    db: Session,
    *,
    template_id: int,
    sections: List[str],
    schema_obj: Dict[str, Any],
    publish: bool,
    user_id: int,
) -> None:
    tpl = db.query(EmrTemplate).filter(EmrTemplate.id == int(template_id)).one_or_none()
    if not tpl:
        return

    max_v = db.query(EmrTemplateVersion).filter(
        EmrTemplateVersion.template_id == int(template_id)
    ).order_by(EmrTemplateVersion.version_no.desc()).first()
    next_no = int(max_v.version_no) + 1 if max_v else 1

    v = EmrTemplateVersion(
        template_id=int(template_id),
        version_no=next_no,
        changelog=f"Seed update v{next_no}",
        sections_json=_dumps(sections),
        schema_json=_dumps(schema_obj),
        created_by_user_id=user_id,
    )
    db.add(v)
    db.flush()

    tpl.active_version_id = int(v.id)
    if publish:
        tpl.status = EmrTemplateStatus.PUBLISHED
        tpl.published_version_id = int(v.id)

    db.commit()


def seed_emr_starter_templates(
    db: Session,
    *,
    excel_path: Optional[str],
    publish: bool,
    update_existing: bool,
    create_missing_departments: bool,
    create_missing_record_types: bool,
    user_id: int = 1,
) -> Dict[str, Any]:
    _ensure_record_types(db, create_missing=create_missing_record_types)
    _ensure_core_sections(db)

    created = 0
    updated = 0
    skipped = 0
    missing_dept = 0
    missing_type = 0

    # --- default starter templates (per dept) ---
    starter_defs = [
        ("OPD_NOTE", "Starter OPD Consultation", "OPD", "Doctor", "Vitals, SOAP, Diagnosis, Medications, Advice"),
        ("IPD_NOTE", "Starter IPD Progress Note", "IPD", "Doctor", "Vitals, SOAP, Diagnosis, Medications, Investigations"),
        ("DISCHARGE_SUMMARY", "Starter Discharge Summary", "IPD", "Doctor", "Diagnosis, Course, Procedures, Discharge Medications, Advice"),
        ("NURSING_NOTE", "Starter Nursing Assessment", "IPD", "Nurse", "Vitals, Nursing Assessment, Pain, Fall risk, Braden, I/O"),
    ]

    # ANC/OT are created for all depts too, but you can restrict later by dept if you prefer
    extra_defs = [
        ("ANC_NOTE", "Starter ANC Visit Note", "OPD/ANC", "Doctor", "ANC Profile, OB History, Vitals, Fetal Assessment, Plan"),
        ("OT_NOTE", "Starter OT Operative / Anaesthesia Note", "OT", "Doctor", "Checklist, Anaesthesia, Procedure, Counts, Post-op orders"),
    ]

    active_depts = db.query(EmrDepartment).filter(EmrDepartment.is_active.is_(True)).order_by(EmrDepartment.display_order.asc()).all()

    # choose default template only if none exists for dept+type
    def has_default(dept_code: str, rt: str) -> bool:
        x = db.query(EmrTemplate.id).filter(
            EmrTemplate.dept_code == dept_code,
            EmrTemplate.record_type_code == rt,
            EmrTemplate.is_default.is_(True),
        ).first()
        return bool(x)

    for d in active_depts:
        dept_code = d.code

        for rt, name, area, who, mandatory in (starter_defs + extra_defs):
            rt = _norm(rt)
            if not _ensure_record_type(db, rt, create_missing_record_types):
                missing_type += 1
                continue

            sections = _sections_for(rt)
            schema_obj = _schema_for(rt, area=area, who=who, mandatory=mandatory)
            is_def = (not has_default(dept_code, rt)) and ("Starter" in name)

            tid = _template_exists(db, dept_code=dept_code, record_type_code=rt, name=name)
            if tid:
                if update_existing:
                    _add_new_version(db, template_id=tid, sections=sections, schema_obj=schema_obj, publish=publish, user_id=user_id)
                    updated += 1
                else:
                    skipped += 1
                continue

            _create_template_with_version(
                db,
                dept_code=dept_code,
                record_type_code=rt,
                name=name,
                description=mandatory,
                sections=sections,
                schema_obj=schema_obj,
                publish=publish,
                is_default=is_def,
                user_id=user_id,
            )
            created += 1

    # --- Excel import (your uploaded sheet list) ---
    if excel_path:
        xrows = _parse_excel(excel_path)

        for xr in xrows:
            rt = _choose_record_type(xr.area, xr.template_name, xr.who)
            rt = _norm(rt)

            if not _ensure_record_type(db, rt, create_missing_record_types):
                missing_type += 1
                continue

            target_depts: List[str] = []

            if xr.dept == "__ALL__":
                # apply to every active dept
                target_depts = [d.code for d in active_depts]
            else:
                dc = _ensure_dept_if_missing(db, xr.dept, create_missing_departments)
                if not dc:
                    missing_dept += 1
                    continue
                target_depts = [dc]

            for dept_code in target_depts:
                name = _safe_name(xr.template_name)
                desc = _safe_desc(xr.mandatory)

                sections = _sections_for(rt)
                schema_obj = _schema_for(rt, area=xr.area, who=xr.who, mandatory=xr.mandatory)

                tid = _template_exists(db, dept_code=dept_code, record_type_code=rt, name=name)
                if tid:
                    if update_existing:
                        _add_new_version(db, template_id=tid, sections=sections, schema_obj=schema_obj, publish=publish, user_id=user_id)
                        updated += 1
                    else:
                        skipped += 1
                    continue

                # default = only if none exists at all for that dept+type
                is_def = not bool(db.query(EmrTemplate.id).filter(
                    EmrTemplate.dept_code == dept_code,
                    EmrTemplate.record_type_code == rt
                ).first())

                _create_template_with_version(
                    db,
                    dept_code=dept_code,
                    record_type_code=rt,
                    name=name,
                    description=desc,
                    sections=sections,
                    schema_obj=schema_obj,
                    publish=publish,
                    is_default=is_def,
                    user_id=user_id,
                )
                created += 1

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "missing_dept": missing_dept,
        "missing_type": missing_type,
        "publish": publish,
        "excel_used": bool(excel_path),
    }


# ----------------------------
# CLI Runner
# ----------------------------
def _make_db_session(db_uri: Optional[str]) -> Session:
    """
    Works with your multi-tenant setup:
    - If db_uri provided: create tenant session using create_tenant_session
    - Else: uses SessionLocal if available
    """
    if db_uri:
        from app.db.session import create_tenant_session
        return create_tenant_session(db_uri)

    try:
        from app.db.session import SessionLocal
        return SessionLocal()
    except Exception as e:
        raise RuntimeError("Unable to create DB session. Provide --db-uri") from e


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db-uri", default=None, help="Tenant DB URI (recommended for multi-tenant)")
    ap.add_argument("--excel", default=None, help="Path to excel (optional). Example: /mnt/data/all department wise casesheet.xlsx")
    ap.add_argument("--publish", action="store_true", help="Publish templates immediately")
    ap.add_argument("--update-existing", action="store_true", help="If template exists, add a new version instead of skipping")
    ap.add_argument("--create-missing-departments", action="store_true", help="Auto-create depts if excel dept not found")
    ap.add_argument("--create-missing-record-types", action="store_true", help="Auto-create missing record types")
    ap.add_argument("--user-id", type=int, default=1, help="Seeder user id")
    args = ap.parse_args()

    db = _make_db_session(args.db_uri)
    try:
        out = seed_emr_starter_templates(
            db,
            excel_path=args.excel,
            publish=bool(args.publish),
            update_existing=bool(args.update_existing),
            create_missing_departments=bool(args.create_missing_departments),
            create_missing_record_types=bool(args.create_missing_record_types),
            user_id=int(args.user_id),
        )
        print("✅ EMR starter template seed done:", out)
    finally:
        db.close()


if __name__ == "__main__":
    main()
